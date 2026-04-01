# experiments/scripts/06b_rebuild_pair_judgements_excel.py
# ============================================================
# Rebuild Step6 aggregation from intermediate JSON files.
#
# What this script does:
# 1) Scan: experiments/runs/<run_id>/analysis/pair_judgements/<model>/<lecture>/<pair_id>/rep_*.json
# 2) Flatten into a long CSV (1 row per judgement key per rep)
# 3) Aggregate into thesis-friendly summary tables
# 4) Export an Excel workbook with multiple sheets
#
# IMPORTANT (user requirement):
# - "Win rate direction" depends on pair orientation.
#   * Non-combo pairs: usually baseline on A, variant on B -> report B win-rate as "focus"
#   * Combo pairs: report A win-rate as "focus"
#   * If B is combo (rare / future): ALSO report A win-rate as "focus" (per user note)
#
# Outputs (default):
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/pair_long_from_json.csv
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/summary_pair_key_overall.csv
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/summary_pair_key_by_lecture.csv
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/summary_model_key_overall.csv
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/thesis_focus_winrate_wide.csv
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/errors.csv
# - experiments/runs/<run_id>/analysis/rebuild_pair_judgements/pair_judgements_thesis.xlsx
# ============================================================

from __future__ import annotations

import argparse
import csv
import json
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from math import sqrt
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Excel (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Alignment = None  # type: ignore
    Font = None  # type: ignore
    get_column_letter = None  # type: ignore


# ----------------------------
# Paths
# ----------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]
RUNS_ROOT = PROJECT_ROOT / "experiments" / "runs"

TEMPLATE_EXPERIMENT_CONFIG_DEFAULT = PROJECT_ROOT / "experiments" / "config" / "experiment_config.json"
RUN_SNAPSHOT_CONFIG = "config/experiment_config.json"


# ----------------------------
# Helpers
# ----------------------------
def now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def safe_float(x: Any, default: float = 0.0) -> float:
    try:
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def mean_std(xs: List[float]) -> Tuple[float, float]:
    if not xs:
        return 0.0, 0.0
    m = sum(xs) / len(xs)
    if len(xs) <= 1:
        return m, 0.0
    v = sum((x - m) ** 2 for x in xs) / (len(xs) - 1)
    return m, sqrt(v)


def sanitize_name(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", str(s)).strip()


def resolve_config_path(override: str, run_dir: Path) -> Path:
    if override.strip():
        p = Path(override)
        if not p.is_absolute():
            p = (Path(".") / p).resolve()
        return p
    snap = run_dir / RUN_SNAPSHOT_CONFIG
    if snap.exists():
        return snap
    return TEMPLATE_EXPERIMENT_CONFIG_DEFAULT


def _ensure_openpyxl() -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not available. Please install openpyxl.")


def _ws_write_table(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    bold = Font(bold=True) if Font else None
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        if bold:
            cell.font = bold
        cell.alignment = Alignment(vertical="top", wrap_text=True) if Alignment else None

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # freeze header row
    ws.freeze_panes = "A2"

    # width auto
    if get_column_letter:
        for i, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
                v = row[0].value
                if v is None:
                    continue
                max_len = max(max_len, min(80, len(str(v))))
            ws.column_dimensions[get_column_letter(i)].width = min(80, max(10, max_len + 2))

    # wrap long columns
    wrap_cols = set()
    for i, h in enumerate(headers, start=1):
        if any(k in h for k in ("reason", "error", "json_path")):
            wrap_cols.add(i)
    if wrap_cols and Alignment:
        for r in range(2, ws.max_row + 1):
            for c in wrap_cols:
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)


# ----------------------------
# Config schema
# ----------------------------
@dataclass(frozen=True)
class CondSpec:
    type: str   # "animation" | "audio"
    level: str  # "intro" | "baseline" | "advanced"
    detail: str # "summary" | "baseline" | "detail"


def load_cond_specs(cfg_path: Path) -> Dict[str, CondSpec]:
    cfg = read_json(cfg_path)
    d = cfg.get("condition_specs", {}) if isinstance(cfg, dict) else {}
    out: Dict[str, CondSpec] = {}
    for k, v in d.items():
        if not isinstance(v, dict):
            continue
        out[str(k)] = CondSpec(
            type=str(v.get("type", "")),
            level=str(v.get("level", "")),
            detail=str(v.get("detail", "")),
        )
    return out


# ----------------------------
# Target / Focus win-rate logic
# ----------------------------
def is_combo(cond_id: str) -> bool:
    return "combo_" in (cond_id or "")


def is_baseline(cond_id: str) -> bool:
    s = (cond_id or "")
    return s.startswith("baseline_") or s in ("baseline_anim", "baseline_audio")


def focus_side_for_pair(A_cond: str, B_cond: str) -> str:
    """
    Return which side ("A" or "B") we should report as the 'focus' win-rate.

    User requirement:
    - If B is combo -> report A win-rate.
    Practical default:
    - If any combo in the pair -> report A win-rate (your current pair design has combo on A)
    - Else if exactly one baseline exists -> report the non-baseline side
    - Else -> report B
    """
    if is_combo(B_cond):
        return "A"  # explicit user requirement
    if is_combo(A_cond) or is_combo(B_cond):
        return "A"

    a_base = is_baseline(A_cond)
    b_base = is_baseline(B_cond)
    if a_base and not b_base:
        return "B"
    if b_base and not a_base:
        return "A"
    return "B"


def winner_to_cond(winner_raw: str, A_cond: str, B_cond: str) -> str:
    if winner_raw == "A":
        return A_cond
    if winner_raw == "B":
        return B_cond
    if winner_raw == "tie":
        return "tie"
    return "unknown"


def expected_target_cond_for_key(
    key: str, A_cond: str, B_cond: str, cond_specs: Dict[str, CondSpec]
) -> str:
    """
    'Target' means: the condition that should be selected if the judgement aligns with the intended direction.
    (Used for 'target_win_rate' = correctness rate)

    Keys seen in your pipeline:
    - summary, detail, intro, advanced
    - modality_audio_fitness (and optionally modality_visual_fitness)
    - goal_deep_listening / goal_overview / goal_quick
    """
    k = str(key or "").strip()
    a = cond_specs.get(A_cond)
    b = cond_specs.get(B_cond)

    def pick(pred) -> str:
        if a and pred(a):
            return A_cond
        if b and pred(b):
            return B_cond
        return ""

    if k == "summary":
        return pick(lambda c: c.detail == "summary")
    if k == "detail":
        return pick(lambda c: c.detail == "detail")
    if k == "intro":
        return pick(lambda c: c.level == "intro")
    if k == "advanced":
        return pick(lambda c: c.level == "advanced")
    if k == "modality_audio_fitness":
        return pick(lambda c: c.type == "audio")
    if k == "modality_visual_fitness":
        return pick(lambda c: c.type == "animation")

    if k in ("goal_deep_listening", "goal_overview", "goal_quick"):
        # by default, "goal_*" aims to choose the combo side if present
        if is_combo(A_cond):
            return A_cond
        if is_combo(B_cond):
            return B_cond
        return ""

    return ""


# ----------------------------
# JSON scanning
# ----------------------------
def scan_rep_json_files(run_dir: Path) -> List[Path]:
    root = run_dir / "analysis" / "pair_judgements"
    if not root.exists():
        return []
    rep_files: List[Path] = []
    for p in root.rglob("rep_*.json"):
        if p.is_file():
            rep_files.append(p)
    rep_files.sort()
    return rep_files


def parse_path_hint(run_dir: Path, rep_json: Path) -> Dict[str, str]:
    """
    Path schema (new):
      .../analysis/pair_judgements/<model>/<lecture>/<pair_id>/rep_01.json
    """
    out = {"model": "", "lecture": "", "pair_id": "", "rep": ""}
    try:
        rel = rep_json.relative_to(run_dir)
        parts = rel.parts
        # find index of ("analysis","pair_judgements")
        for i in range(len(parts) - 2):
            if parts[i] == "analysis" and parts[i + 1] == "pair_judgements":
                # model, lecture, pair_id, repfile
                out["model"] = parts[i + 2] if i + 2 < len(parts) else ""
                out["lecture"] = parts[i + 3] if i + 3 < len(parts) else ""
                out["pair_id"] = parts[i + 4] if i + 4 < len(parts) else ""
                fname = parts[i + 5] if i + 5 < len(parts) else rep_json.name
                m = re.search(r"rep_(\d+)\.json$", fname)
                out["rep"] = m.group(1) if m else ""
                break
    except Exception:
        pass
    return out


# ----------------------------
# Aggregation
# ----------------------------
def aggregate_group(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Given long rows in the same group, compute counts & win rates.
    Denominator excludes tie/unknown for win rates.
    """
    winners = [str(r.get("winner_raw", "")).strip() for r in rows]
    cnt = Counter(winners)

    confs = [safe_float(r.get("confidence", ""), 0.0) for r in rows]
    m, s = mean_std(confs)

    # focus
    focus_hits = 0
    focus_valid = 0
    target_hits = 0
    target_valid = 0

    for r in rows:
        wr = str(r.get("winner_raw", "")).strip()
        if wr not in ("A", "B"):
            continue

        # focus
        fv = r.get("is_focus", "")
        if fv in (0, 1):
            focus_valid += 1
            if fv == 1:
                focus_hits += 1

        # target
        tv = r.get("is_target", "")
        if tv in (0, 1):
            target_valid += 1
            if tv == 1:
                target_hits += 1

    majority = ""
    top = cnt.most_common()
    if top:
        if len(top) >= 2 and top[0][1] == top[1][1]:
            majority = "tie"
        else:
            majority = top[0][0]

    return {
        "n_trials": len(rows),
        "A": cnt.get("A", 0),
        "B": cnt.get("B", 0),
        "tie": cnt.get("tie", 0),
        "other": sum(cnt.values()) - cnt.get("A", 0) - cnt.get("B", 0) - cnt.get("tie", 0),
        "mean_confidence": m,
        "std_confidence": s,
        "majority_winner": majority,
        "focus_n": focus_valid,
        "focus_win_rate": (focus_hits / focus_valid) if focus_valid else "",
        "target_n": target_valid,
        "target_win_rate": (target_hits / target_valid) if target_valid else "",
    }


def make_wide_focus_table(summary_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Pivot summary_pair_key_overall into wide format:
      row = pair_id + key
      columns = focus_win_rate_<model>
    """
    models = sorted({str(r.get("model", "")) for r in summary_rows if str(r.get("model", ""))})
    index = sorted({(str(r.get("pair_id", "")), str(r.get("key", ""))) for r in summary_rows})

    # lookup
    lut: Dict[Tuple[str, str, str], Any] = {}
    meta_lut: Dict[Tuple[str, str, str], Any] = {}
    for r in summary_rows:
        k = (str(r.get("pair_id", "")), str(r.get("key", "")), str(r.get("model", "")))
        lut[k] = r.get("focus_win_rate", "")
        meta_lut[k] = r

    out: List[Dict[str, Any]] = []
    for pair_id, key in index:
        row: Dict[str, Any] = {"pair_id": pair_id, "key": key}
        # optionally attach A/B/focus
        # take first model's meta if available
        for m in models:
            meta = meta_lut.get((pair_id, key, m))
            if meta:
                row.setdefault("A_cond", meta.get("A_cond", ""))
                row.setdefault("B_cond", meta.get("B_cond", ""))
                row.setdefault("focus_side", meta.get("focus_side", ""))
                row.setdefault("focus_cond", meta.get("focus_cond", ""))
                break

        for m in models:
            row[f"focus_win_rate_{m}"] = lut.get((pair_id, key, m), "")
        out.append(row)
    return out


# ----------------------------
# Main rebuild
# ----------------------------
def rebuild(run_id: str, cfg_path: Path, out_dir: Path, excel_path: Path) -> None:
    run_dir = RUNS_ROOT / run_id
    if not run_dir.exists():
        raise FileNotFoundError(f"run_dir not found: {run_dir}")

    cond_specs = load_cond_specs(cfg_path)

    rep_files = scan_rep_json_files(run_dir)
    if not rep_files:
        raise FileNotFoundError(f"No rep_*.json found under: {run_dir/'analysis/pair_judgements'}")

    long_rows: List[Dict[str, Any]] = []
    err_rows: List[Dict[str, Any]] = []

    for rep_json in rep_files:
        hint = parse_path_hint(run_dir, rep_json)
        model = hint.get("model", "")
        lecture = hint.get("lecture", "")
        pair_id = hint.get("pair_id", "")
        rep = hint.get("rep", "")

        try:
            payload = read_json(rep_json)
        except Exception as e:
            err_rows.append({
                "run_id": run_id,
                "model": model,
                "lecture": lecture,
                "pair_id": pair_id,
                "rep": rep,
                "error": f"JSON_READ_FAILED: {type(e).__name__}: {e}",
                "json_path": str(rep_json),
            })
            continue

        A_cond = str(payload.get("A", "")).strip()
        B_cond = str(payload.get("B", "")).strip()

        judgements = payload.get("judgements")
        if not isinstance(judgements, dict):
            err_rows.append({
                "run_id": run_id,
                "model": model,
                "lecture": lecture or str(payload.get("lecture", "")),
                "pair_id": pair_id or str(payload.get("pair_id", "")),
                "rep": rep,
                "error": "JUDGEMENTS_MISSING_OR_NOT_DICT",
                "json_path": str(rep_json),
            })
            continue

        # timestamp (best effort)
        ts = ""
        meta = payload.get("_meta", {})
        if isinstance(meta, dict):
            ts = str(meta.get("timestamp", "")).strip()

        focus_side = focus_side_for_pair(A_cond, B_cond)
        focus_cond = A_cond if focus_side == "A" else B_cond

        for key, j in judgements.items():
            if not isinstance(j, dict):
                continue
            key = str(key)
            winner_raw = str(j.get("winner", "")).strip()
            conf = safe_float(j.get("confidence", ""), 0.0)
            reason = str(j.get("reason", "")).strip()

            winner_cond = winner_to_cond(winner_raw, A_cond, B_cond)

            # focus correctness
            is_focus: Any = ""
            if winner_raw in ("A", "B"):
                is_focus = 1 if winner_raw == focus_side else 0

            # target correctness (config-based)
            target_cond = expected_target_cond_for_key(key, A_cond, B_cond, cond_specs)
            is_target: Any = ""
            if target_cond and winner_cond not in ("tie", "unknown"):
                is_target = 1 if winner_cond == target_cond else 0

            long_rows.append({
                "run_id": run_id,
                "model": model,
                "lecture": lecture or str(payload.get("lecture", "")),
                "pair_id": pair_id or str(payload.get("pair_id", "")),
                "rep": rep,
                "key": key,
                "A_cond": A_cond,
                "B_cond": B_cond,
                "winner_raw": winner_raw,
                "winner_cond": winner_cond,
                "confidence": conf,
                "reason": reason,
                "focus_side": focus_side,
                "focus_cond": focus_cond,
                "is_focus": is_focus,
                "target_cond": target_cond,
                "is_target": is_target,
                "json_path": str(rep_json),
                "timestamp": ts,
            })

    # ----------------------------
    # Write CSVs
    # ----------------------------
    out_dir.mkdir(parents=True, exist_ok=True)

    def write_csv(path: Path, headers: List[str], rows: List[Dict[str, Any]]) -> None:
        with path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, "") for h in headers})

    long_headers = [
        "run_id", "model", "lecture", "pair_id", "rep",
        "key", "A_cond", "B_cond",
        "winner_raw", "winner_cond",
        "confidence", "reason",
        "focus_side", "focus_cond", "is_focus",
        "target_cond", "is_target",
        "json_path", "timestamp",
    ]
    write_csv(out_dir / "pair_long_from_json.csv", long_headers, long_rows)

    err_headers = ["run_id", "model", "lecture", "pair_id", "rep", "error", "json_path"]
    write_csv(out_dir / "errors.csv", err_headers, err_rows)

    # ----------------------------
    # Aggregations
    # ----------------------------
    # (1) model × pair × key (overall across lectures)
    by_mpk: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in long_rows:
        by_mpk[(str(r["model"]), str(r["pair_id"]), str(r["key"]))].append(r)

    summary_pair_key_overall: List[Dict[str, Any]] = []
    for (model, pair_id, key), rows in sorted(by_mpk.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        agg = aggregate_group(rows)
        # attach stable A/B/focus info from first row
        base = rows[0]
        summary_pair_key_overall.append({
            "model": model,
            "pair_id": pair_id,
            "key": key,
            "A_cond": base.get("A_cond", ""),
            "B_cond": base.get("B_cond", ""),
            "focus_side": base.get("focus_side", ""),
            "focus_cond": base.get("focus_cond", ""),
            "target_cond": base.get("target_cond", ""),
            **agg,
        })

    # (2) model × lecture × pair × key
    by_mlpk: Dict[Tuple[str, str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in long_rows:
        by_mlpk[(str(r["model"]), str(r["lecture"]), str(r["pair_id"]), str(r["key"]))].append(r)

    summary_pair_key_by_lecture: List[Dict[str, Any]] = []
    for (model, lecture, pair_id, key), rows in sorted(by_mlpk.items(), key=lambda x: (x[0][0], x[0][1], x[0][2], x[0][3])):
        agg = aggregate_group(rows)
        base = rows[0]
        summary_pair_key_by_lecture.append({
            "model": model,
            "lecture": lecture,
            "pair_id": pair_id,
            "key": key,
            "A_cond": base.get("A_cond", ""),
            "B_cond": base.get("B_cond", ""),
            "focus_side": base.get("focus_side", ""),
            "focus_cond": base.get("focus_cond", ""),
            "target_cond": base.get("target_cond", ""),
            **agg,
        })

    # (3) model × key (overall)
    by_mk: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in long_rows:
        by_mk[(str(r["model"]), str(r["key"]))].append(r)

    summary_model_key_overall: List[Dict[str, Any]] = []
    for (model, key), rows in sorted(by_mk.items(), key=lambda x: (x[0][0], x[0][1])):
        agg = aggregate_group(rows)
        summary_model_key_overall.append({
            "model": model,
            "key": key,
            **agg,
        })

    # thesis pivot (wide)
    thesis_wide = make_wide_focus_table(summary_pair_key_overall)

    # write summary CSVs
    spk_headers = [
        "model", "pair_id", "key",
        "A_cond", "B_cond",
        "focus_side", "focus_cond",
        "target_cond",
        "n_trials", "A", "B", "tie", "other",
        "mean_confidence", "std_confidence", "majority_winner",
        "focus_n", "focus_win_rate",
        "target_n", "target_win_rate",
    ]
    write_csv(out_dir / "summary_pair_key_overall.csv", spk_headers, summary_pair_key_overall)

    spkL_headers = [
        "model", "lecture", "pair_id", "key",
        "A_cond", "B_cond",
        "focus_side", "focus_cond",
        "target_cond",
        "n_trials", "A", "B", "tie", "other",
        "mean_confidence", "std_confidence", "majority_winner",
        "focus_n", "focus_win_rate",
        "target_n", "target_win_rate",
    ]
    write_csv(out_dir / "summary_pair_key_by_lecture.csv", spkL_headers, summary_pair_key_by_lecture)

    smk_headers = [
        "model", "key",
        "n_trials", "A", "B", "tie", "other",
        "mean_confidence", "std_confidence", "majority_winner",
        "focus_n", "focus_win_rate",
        "target_n", "target_win_rate",
    ]
    write_csv(out_dir / "summary_model_key_overall.csv", smk_headers, summary_model_key_overall)

    wide_headers = list(thesis_wide[0].keys()) if thesis_wide else ["pair_id", "key"]
    write_csv(out_dir / "thesis_focus_winrate_wide.csv", wide_headers, thesis_wide)

    # ----------------------------
    # Excel export
    # ----------------------------
    _ensure_openpyxl()
    wb = Workbook()
    wb.remove(wb.active)

    ws_info = wb.create_sheet("info")
    _ws_write_table(ws_info, ["key", "value"], [
        {"key": "run_id", "value": run_id},
        {"key": "run_dir", "value": str(run_dir)},
        {"key": "config_used", "value": str(cfg_path)},
        {"key": "generated_at", "value": now_iso()},
        {"key": "n_rep_files", "value": len(rep_files)},
        {"key": "n_long_rows", "value": len(long_rows)},
        {"key": "n_errors", "value": len(err_rows)},
        {"key": "focus_rule", "value": "combo -> A (if B is combo, still A). else non-baseline; else B"},
    ])

    ws_long = wb.create_sheet("pair_long")
    _ws_write_table(ws_long, long_headers, long_rows)

    ws_spk = wb.create_sheet("summary_pair_key")
    _ws_write_table(ws_spk, spk_headers, summary_pair_key_overall)

    ws_spkL = wb.create_sheet("summary_pair_key_by_lecture")
    _ws_write_table(ws_spkL, spkL_headers, summary_pair_key_by_lecture)

    ws_smk = wb.create_sheet("summary_model_key")
    _ws_write_table(ws_smk, smk_headers, summary_model_key_overall)

    ws_wide = wb.create_sheet("thesis_focus_winrate_wide")
    _ws_write_table(ws_wide, wide_headers, thesis_wide)

    ws_err = wb.create_sheet("errors")
    _ws_write_table(ws_err, err_headers, err_rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-id", required=True, help="experiments/runs/<run_id>")

    ap.add_argument("--config", default="", help="override config path (default: run snapshot then template)")
    ap.add_argument("--out-dir", default="", help="override output dir (default: <run>/analysis/rebuild_pair_judgements)")
    ap.add_argument("--excel-path", default="", help="override excel path (default: <out_dir>/pair_judgements_thesis.xlsx)")

    args = ap.parse_args()

    run_dir = RUNS_ROOT / args.run_id
    if not run_dir.exists():
        raise FileNotFoundError(f"run_dir not found: {run_dir}")

    cfg_path = resolve_config_path(args.config, run_dir)
    if not cfg_path.exists():
        raise FileNotFoundError(f"config not found: {cfg_path}")

    out_dir = Path(args.out_dir).resolve() if args.out_dir.strip() else (run_dir / "analysis" / "rebuild_pair_judgements")
    excel_path = Path(args.excel_path).resolve() if args.excel_path.strip() else (out_dir / "pair_judgements_thesis.xlsx")

    rebuild(args.run_id, cfg_path, out_dir, excel_path)
    print(f"[06b] done.")
    print(f"[06b] out_dir : {out_dir}")
    print(f"[06b] excel   : {excel_path}")


if __name__ == "__main__":
    main()
