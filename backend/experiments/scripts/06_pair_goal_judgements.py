# experiments/scripts/06_pair_goal_judgements.py
# ============================================================
# Step6: LLMによるペア比較判定（summary/intro/detail/advanced/modality + goal）
#
# 変更点（2026-01）
# - 判定モデルを固定3種（デフォルト）：GPT3.5 / GPT4o / GPT5.2
# - 同一の「一連の質問（keys）」を N 回繰り返して集計（デフォルト N=10）
# - Excel（.xlsx）に「モデルごと × 質問（key）ごと」で平均・勝率を集計
#
# 【参照仕様（重要）】
# - pairs / judge_models / condition_specs は config を参照：
#     (優先) experiments/runs/<run_id>/config/experiment_config.json
#     (fallback) experiments/config/experiment_config.json
# - 生成物の実体（各condの出力場所）は run_id の Step2 manifest から解決：
#     experiments/runs/<run_id>/generation/manifest_step02.jsonl
#
# 【keys の仕様（ユーザ指定）】
# - baseline vs summary        -> keys=["summary"]
# - baseline vs intro          -> keys=["intro"]
# - baseline vs detail         -> keys=["detail"]
# - baseline vs advanced       -> keys=["advanced"]
# - type が animation vs audio -> keys に "modality" を追加（2項目を判定）
#
# 【追加ペアの専用質問（ユーザ指定）】
# - combo_adv_detail_anim vs baseline_anim
#     -> keys に "goal_deep_listening" を追加
# - combo_intro_summary_audio vs baseline_audio
#     -> keys に "goal_overview" を追加
# - combo_intro_summary_audio vs baseline_anim
#     -> keys に "goal_quick" を追加（+ modality も発生し得る）
#
# 【出力】
# - experiments/runs/<run_id>/analysis/pair_judgements.csv
#     * 1行 = 1モデル × 1ペア × 1講義 × 1反復(rep)
# - experiments/runs/<run_id>/analysis/pair_judgements/<model>/<lecture>/<pair_id>/rep_01.json ... rep_10.json
# - experiments/runs/<run_id>/analysis/pair_judgements/<model>/<lecture>/<pair_id>/aggregate.json
# - experiments/runs/<run_id>/analysis/pair_judgements.xlsx
# ============================================================

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from math import sqrt
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openai import OpenAI

# Excel (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Alignment = None  # type: ignore
    Font = None  # type: ignore
    get_column_letter = None  # type: ignore


# ----------------------------
# Paths
# ----------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]
RUNS_ROOT = PROJECT_ROOT / "experiments" / "runs"

TEMPLATE_EXPERIMENT_CONFIG_DEFAULT = PROJECT_ROOT / "experiments" / "config" / "experiment_config.json"
RUN_SNAPSHOT_CONFIG = "config/experiment_config.json"
APIKEY_DEFAULT = PROJECT_ROOT / "apikey.txt"


# ----------------------------
# Helpers
# ----------------------------
def now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())


def read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, obj: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def read_text_best_effort(path: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp932"):
        try:
            return path.read_text(encoding=enc)
        except Exception:
            continue
    return path.read_text(errors="ignore")


def sanitize_name(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", str(s)).strip()



def _relpath_from_run(p: Path, run_dir: Path) -> Path:
    """Return path relative to run_dir if possible, else original."""
    try:
        return p.relative_to(run_dir)
    except Exception:
        return p


def resolve_json_path_for_row(row: Dict[str, Any], run_dir: Path) -> Path:
    """Robustly resolve JSON path even if CSV contains stale absolute paths."""
    # 1) prefer json_relpath (portable)
    rel = str(row.get("json_relpath", "")).strip()
    if rel:
        cand = run_dir / Path(rel)
        if cand.exists():
            return cand

    # 2) try json_path as-is
    jp_s = str(row.get("json_path", "")).strip()
    if jp_s:
        p = Path(jp_s)
        if p.exists():
            return p

        # 2b) rebase from .../analysis/pair_judgements/... tail (moved run folder)
        parts = list(p.parts)
        for i in range(len(parts) - 1):
            if str(parts[i]).lower() == "analysis" and str(parts[i + 1]).lower() == "pair_judgements":
                tail = Path(*parts[i:])  # analysis/...
                cand = run_dir / tail
                if cand.exists():
                    return cand

    # 3) reconstruct from schema (new/old)
    model = str(row.get("model", "")).strip()
    lecture = sanitize_name(str(row.get("lecture", "")).strip())
    pair_id = str(row.get("pair_id", "")).strip()
    rep_s = str(row.get("rep", "")).strip()

    base = run_dir / "analysis" / "pair_judgements" / model / lecture
    if rep_s:
        try:
            rep_i = int(rep_s)
            cand = base / pair_id / f"rep_{rep_i:02d}.json"
            if cand.exists():
                return cand
        except Exception:
            pass

    cand_old = base / f"{pair_id}.json"
    return cand_old


def safe_float(x: Any, default: float = 0.0) -> float:
    try:
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def mean_std(xs: List[float]) -> Tuple[float, float]:
    if not xs:
        return 0.0, 0.0
    m = sum(xs) / len(xs)
    if len(xs) <= 1:
        return m, 0.0
    v = sum((x - m) ** 2 for x in xs) / (len(xs) - 1)
    return m, sqrt(v)


def resolve_config_path(override: str, run_dir: Path) -> Path:
    if override.strip():
        p = Path(override)
        if not p.is_absolute():
            p = (Path(".") / p).resolve()
        return p
    snap = run_dir / RUN_SNAPSHOT_CONFIG
    if snap.exists():
        return snap
    return TEMPLATE_EXPERIMENT_CONFIG_DEFAULT


# ----------------------------
# Model presets / normalization
# ----------------------------
DEFAULT_FIXED_MODELS = ["gpt-4o", "gpt-4o-mini", "gpt-5", "gpt-5.2"]

# "gpt-3.5-turbo", 
def normalize_model_id(name: str) -> str:
    """
    Accept loose user names like 'GPT3.5', 'GPT4o', 'GPT5.2' and map to API model IDs when possible.
    If unknown, return as-is.
    """
    raw = str(name or "").strip()
    if not raw:
        return raw
    t = raw.lower().strip()
    key = re.sub(r"[\s_\-]+", "", t)

    # GPT-3.5
    if key in ("gpt35", "gpt3.5", "gpt35turbo", "gpt35t"):
        return "gpt-3.5-turbo"

    # GPT-4o
    if key in ("gpt4o", "gpt4.0o", "gpt4omni"):
        return "gpt-4o"

    # GPT-4o mini
    if key in ("gpt4omini", "gpt4omin", "4omini", "gpt4omini"):
        return "gpt-4o-mini"

    # GPT-5.2
    if key in ("gpt52", "gpt5.2", "gpt5_2"):
        return "gpt-5.2"

    return raw


# ============================================================
# GPT-5 対応 JSON 抽出（堅牢版）
# ============================================================
def extract_first_json_object(s: str) -> Optional[Dict[str, Any]]:
    if not s:
        return None

    text = s.strip()

    # ① 素の JSON をまず試す（GPT-5 対応）
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass

    # ② ```json ... ``` ブロック
    m = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text, flags=re.IGNORECASE)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            pass

    # ③ 最外殻の { ... } を貪欲に拾う
    m2 = re.search(r"(\{[\s\S]*\})", text)
    if m2:
        try:
            return json.loads(m2.group(1))
        except Exception:
            pass

    return None


# ----------------------------
# API Key
# ----------------------------
def load_api_key_from_file(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"apikey file not found: {path}")
    key = path.read_text(encoding="utf-8", errors="ignore").strip()
    if not key:
        raise RuntimeError(f"apikey file is empty: {path}")
    return key


def get_api_key(apikey_path: Optional[Path]) -> str:
    env = os.environ.get("OPENAI_API_KEY", "").strip()
    if env:
        return env
    path = apikey_path if apikey_path else APIKEY_DEFAULT
    if not path.is_absolute():
        path = (PROJECT_ROOT / path).resolve()
    return load_api_key_from_file(path)


def create_client(apikey_path: Optional[Path]) -> OpenAI:
    return OpenAI(api_key=get_api_key(apikey_path))


# ----------------------------
# Step2 manifest
# ----------------------------
@dataclass(frozen=True)
class Step2Row:
    lecture: str
    cond_id: str
    status: str
    dest: str

    @staticmethod
    def from_dict(d: dict) -> "Step2Row":
        return Step2Row(
            lecture=str(d.get("lecture") or ""),
            cond_id=str(d.get("cond_id") or ""),
            status=str(d.get("status") or ""),
            dest=str(d.get("dest") or ""),
        )


def load_manifest_jsonl(path: Path) -> List[Step2Row]:
    rows: List[Step2Row] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(Step2Row.from_dict(json.loads(line)))
    return rows


def resolve_dest_path(dest_str: str, run_dir: Path) -> Path:
    p = Path(dest_str)

    # 1. 絶対パスならそのまま
    if p.is_absolute():
        return p

    # 2. PROJECT_ROOT 基準で解決（Step2 仕様）
    cand = (PROJECT_ROOT / p).resolve()
    if cand.exists():
        return cand

    # 3. 念のため run_dir 基準も試す（保険）
    cand2 = (run_dir / p).resolve()
    if cand2.exists():
        return cand2

    # 4. 最後にそのまま返す（エラー用）
    return cand


def detect_mode(output_root: Path) -> str:
    lt = output_root / "lecture_outputs" / "lecture_texts"
    if (lt / "all_explanations.txt").exists():
        return "animation"
    if (lt / "audio_only").exists():
        return "audio"
    return "unknown"


def pick_script_path(output_root: Path, mode: str) -> Tuple[Optional[Path], str]:
    lt = output_root / "lecture_outputs" / "lecture_texts"
    if mode == "animation":
        p = lt / "all_explanations.txt"
        if p.exists():
            return p, "animation:lecture_texts/all_explanations.txt"
    if mode == "audio":
        ao = lt / "audio_only"
        for name in ("lecture_script_stitched.txt", "lecture_script.txt"):
            p = ao / name
            if p.exists():
                return p, f"audio:lecture_texts/audio_only/{name}"
    return None, "not_found"


def build_index(rows: List[Step2Row], run_dir: Path) -> Dict[Tuple[str, str], Path]:
    idx: Dict[Tuple[str, str], Path] = {}
    for r in rows:
        if r.status in ("moved", "skip"):
            idx[(r.lecture, r.cond_id)] = resolve_dest_path(r.dest, run_dir)
    return idx


# ----------------------------
# Template config
# ----------------------------
@dataclass(frozen=True)
class PairSpec:
    pair_id: str
    A: str
    B: str


@dataclass(frozen=True)
class CondSpec:
    type: str
    level: str
    detail: str


def load_template_config(path: Path) -> Tuple[List[str], List[PairSpec], Dict[str, CondSpec]]:
    cfg = read_json(path)
    judge_models = list(cfg.get("judge_models", []))
    pairs = [PairSpec(**p) for p in cfg.get("pairs", [])]
    cond_specs = {k: CondSpec(**v) for k, v in cfg.get("condition_specs", {}).items()}
    return judge_models, pairs, cond_specs


# ----------------------------
# Decide which dimensions to judge for a given pair (ユーザ仕様版)
# ----------------------------
def _is_baseline_vs_summary(a: CondSpec, b: CondSpec) -> bool:
    s = {a.detail, b.detail}
    return ("baseline" in s) and ("summary" in s)


def _is_baseline_vs_detail(a: CondSpec, b: CondSpec) -> bool:
    s = {a.detail, b.detail}
    return ("baseline" in s) and ("detail" in s)


def _is_baseline_vs_intro(a: CondSpec, b: CondSpec) -> bool:
    s = {a.level, b.level}
    return ("baseline" in s) and ("intro" in s)


def _is_baseline_vs_advanced(a: CondSpec, b: CondSpec) -> bool:
    s = {a.level, b.level}
    return ("baseline" in s) and ("advanced" in s)


def needed_judgements(pair_id: str, A_cond: str, B_cond: str, a: CondSpec, b: CondSpec) -> List[str]:
    """
    返り値：judgement keys
      - "summary" : どちらが要約（大枠・簡潔）か（baseline vs summary のとき）
      - "detail"  : どちらが詳細（掘り下げ）か（baseline vs detail のとき）
      - "intro"   : どちらが入門向けか（baseline vs intro のとき）
      - "advanced": どちらが発展的か（baseline vs advanced のとき）
      - "modality": audio_fitness / visual_fitness（typeが違うとき）
      - "goal_deep_listening": じっくり聞きたいならどちらが良いか（特定ペア）
      - "goal_overview": 大枠を知りたいならどちらが良いか（特定ペア）
      - "goal_quick": 簡単に知りたいならどちらが良いか（特定ペア）
    """
    keys: List[str] = []

    # --- detail axis ---
    if a.detail != b.detail:
        if _is_baseline_vs_summary(a, b):
            keys.append("summary")
        elif _is_baseline_vs_detail(a, b):
            keys.append("detail")
        else:
            # comboなど（例：summary vs detail）の場合は、まず「detail」で統一
            keys.append("detail")

    # --- level axis ---
    if a.level != b.level:
        if _is_baseline_vs_intro(a, b):
            keys.append("intro")
        elif _is_baseline_vs_advanced(a, b):
            keys.append("advanced")
        else:
            keys.append("advanced")

    # --- modality axis ---
    if a.type != b.type:
        keys.append("modality")

    # --- special goal questions (pair-specific) ---
    if {A_cond, B_cond} == {"combo_adv_detail_anim", "baseline_anim"}:
        keys.append("goal_deep_listening")
    if {A_cond, B_cond} == {"combo_intro_summary_audio", "baseline_audio"}:
        keys.append("goal_overview")
    if {A_cond, B_cond} == {"combo_intro_summary_audio", "baseline_anim"}:
        keys.append("goal_quick")

    # 重複除去（順序保持）
    seen = set()
    uniq: List[str] = []
    for k in keys:
        if k not in seen:
            uniq.append(k)
            seen.add(k)
    return uniq


def expand_out_keys(keys: List[str]) -> List[str]:
    """Expand high-level keys to actual JSON judgement keys."""
    out_keys: List[str] = []
    if "summary" in keys:
        out_keys.append("summary")
    if "detail" in keys:
        out_keys.append("detail")
    if "intro" in keys:
        out_keys.append("intro")
    if "advanced" in keys:
        out_keys.append("advanced")
    if "modality" in keys:
        out_keys.append("modality_audio_fitness")
    if "goal_deep_listening" in keys:
        out_keys.append("goal_deep_listening")
    if "goal_overview" in keys:
        out_keys.append("goal_overview")
    if "goal_quick" in keys:
        out_keys.append("goal_quick")
    return out_keys


# ----------------------------
# Prompt building
# ----------------------------
def build_pair_prompt(
    lecture_key: str,
    pair_id: str,
    A_cond: str,
    B_cond: str,
    A_text: str,
    B_text: str,
    keys: List[str],
) -> Tuple[str, str]:
    """
    JSON only output.
    winner: "A" or "B" or "tie"
    confidence: 0.0-1.0
    """
    system = (
        "あなたは講義テキストの比較評価者です。\n"
        "2つの講義テキスト(A,B)を読み比べ、指定された観点についてどちらが該当するかを判断してください。\n"
        "重要: JSONオブジェクトのみを出力し、余計な文章は一切出力しないでください。\n"
        "重要: judgements 内の各項目は必ず winner/confidence/reason を埋めてください。reason は必ず日本語で1〜3文、空欄禁止です。\n"
        "重要: confidence は 0.01〜1.0 の数値で必ず入力してください（0.0は禁止）。\n"
        "以下，難易度軸と詳細度軸を定義します。\n"
        "難易度軸：\n"
        "intro（入門）: 小学生が聞いても雰囲気がわかるような内容とし，専門用語は用いない．身近な具体例やたとえ話だけで，興味を失わないようにやさしく説明する．\n"
        "baseline（標準）: 基礎概念を平易に説明し，比喩や具体例を多く用いる．専門用語の使用は最小限とし，使用する場合は丁寧に説明する．\n"
        "advanced（発展）: 重要な内容であれば，基本的な専門用語や理論的・抽象的な説明を許容する．スライドの主題から大きく外れない範囲で，関連する応用例や周辺トピックを横方向に広げて紹介する．\n"
        "詳細度軸：\n"
        "summary（要約）: スライドの要点のみを端的に述べる．不要な背景説明や寄り道は行わない．\n"
        "baseline（標準）: 要点に加えて，理解に必要な補足や短い具体例を適切に追加する．ただし冗長にはしない．\n"
        "detail（詳細）: 概念間の関係，背景となる考え方，直感的理解を助ける説明を含めて詳しく解説する．ただしスライドの主題から脱線しない．\n"
    )

    items: List[str] = []
    out_keys: List[str] = expand_out_keys(keys)

    if "summary" in keys:
        items.append("- summary: どちらが要約か")

    if "detail" in keys:
        items.append("- detail: どちらが詳細か")

    if "intro" in keys:
        items.append("- intro: どちらが入門向けか")

    if "advanced" in keys:
        items.append("- advanced: どちらが発展的か")

    if "modality" in keys:
        items.append("- modality_audio_fitness: どちらが音声のみでも理解しやすいか（図やスライド参照が少ない／文章だけで自己完結）")

    if "goal_deep_listening" in keys:
        items.append("- goal_deep_listening: 「じっくり講義を聞きたい」場合どちらが良いか（丁寧、網羅的、深い理解に向く）")

    if "goal_overview" in keys:
        items.append("- goal_overview: 「講義内容の大枠を知りたい」場合どちらが良いか（全体像、要点、短時間把握）")

    if "goal_quick" in keys:
        items.append("- goal_quick: 「講義内容を簡単に知りたい」場合どちらが良いか（負担が軽い、簡潔、すぐ分かる）")

    judgement_lines = []
    for k in out_keys:
        # placeholders are intentionally non-empty to discourage empty outputs
        judgement_lines.append(
            f'    "{k}": {{"winner": "A|B|tie", "confidence": 0.55, "reason": "日本語で理由（1〜3文、空欄禁止）"}}'
        )
    judgements_block = ",\n".join(judgement_lines) if judgement_lines else ""

    user = (
        f"【講義】{lecture_key}\n"
        f"【pair_id】{pair_id}\n"
        f"【A】{A_cond}\n"
        f"【B】{B_cond}\n\n"
        "【判定する観点】\n"
        + "\n".join(items)
        + "\n\n"
        "【Aテキスト】\n-----\n"
        + A_text
        + "\n-----\n\n"
        "【Bテキスト】\n-----\n"
        + B_text
        + "\n-----\n\n"
        "出力JSON形式（厳守）：\n"
        "{\n"
        f'  "lecture": "{lecture_key}",\n'
        f'  "pair_id": "{pair_id}",\n'
        f'  "A": "{A_cond}",\n'
        f'  "B": "{B_cond}",\n'
        '  "judgements": {\n'
        + judgements_block
        + "\n  },\n"
        '  "notes": ""\n'
        "}\n"
        "※ confidence は 0.0〜1.0（主観確信度）\n"
        "※ reason は必ず日本語で1〜3文。空欄禁止。\n"
        "※ JSONのみ出力\n"
    )
    return system, user


def _has_japanese_chars(s: str) -> bool:
    """Rough check: contains Hiragana/Katakana/Kanji."""
    return bool(re.search(r"[\u3040-\u30ff\u4e00-\u9fff]", s or ""))


def _validate_parsed_judgements(parsed: Dict[str, Any], required_out_keys: List[str]) -> Tuple[bool, str]:
    """Validate that all required judgement items are present and filled."""
    if not isinstance(parsed, dict):
        return False, "PARSED_NOT_DICT"

    js = parsed.get("judgements")
    if not isinstance(js, dict):
        return False, "JUDGEMENTS_MISSING"

    for k in required_out_keys:
        j = js.get(k)
        if not isinstance(j, dict):
            return False, f"MISSING_KEY:{k}"

        w = str(j.get("winner", "")).strip()
        if w not in ("A", "B", "tie"):
            return False, f"BAD_WINNER:{k}:{w}"

        conf_raw = j.get("confidence", None)
        try:
            conf = float(conf_raw)
        except Exception:
            return False, f"BAD_CONFIDENCE_TYPE:{k}"
        # enforce non-zero to avoid placeholder 0.0 sticking
        if not (0.01 <= conf <= 1.0):
            return False, f"BAD_CONFIDENCE_RANGE:{k}:{conf}"

        reason = str(j.get("reason", "")).strip()
        if not reason:
            return False, f"EMPTY_REASON:{k}"
        if "日本語で理由" in reason:
            return False, f"REASON_PLACEHOLDER:{k}"
        if not _has_japanese_chars(reason):
            return False, f"REASON_NOT_JA:{k}"

    return True, "OK"


def call_llm(
    client: OpenAI,
    model: str,
    system: str,
    user: str,
    required_out_keys: Optional[List[str]] = None,
    max_retries: int = 1,
) -> Tuple[Optional[Dict[str, Any]], Dict[str, Any]]:
    """Call LLM and enforce non-empty Japanese reasons + non-zero confidence.

    If the parsed JSON is missing required fields, retry once with a stricter instruction.
    """
    t0 = time.time()
    model_used = model

    def _do_call(user_text: str) -> Tuple[Optional[Any], Optional[str]]:
        nonlocal model_used
        try:
            resp = client.responses.create(
                model=model_used,
                input=[
                    {"role": "system", "content": [{"type": "input_text", "text": system}]},
                    {"role": "user", "content": [{"type": "input_text", "text": user_text}]},
                ],
            )
            return resp, None
        except Exception as e:
            msg = f"{type(e).__name__}: {e}"
            # Fallback: some environments expose GPT-5.2 as "gpt-5" (or vice versa).
            if model_used == "gpt-5.2":
                try_model = "gpt-5"
                try:
                    resp = client.responses.create(
                        model=try_model,
                        input=[
                            {"role": "system", "content": [{"type": "input_text", "text": system}]},
                            {"role": "user", "content": [{"type": "input_text", "text": user_text}]},
                        ],
                    )
                    model_used = try_model
                    return resp, None
                except Exception as e2:
                    return None, f"responses.create failed: {msg} | fallback {try_model} failed: {type(e2).__name__}: {e2}"
            return None, f"responses.create failed: {msg}"

    def _extract_text(resp: Any) -> str:
        try:
            t = resp.output_text  # type: ignore
        except Exception:
            t = ""
        return t or str(resp)

    attempts = 0
    last_raw = ""
    last_err = ""
    parsed: Optional[Dict[str, Any]] = None
    validation_msg = ""

    while attempts <= max_retries:
        attempts += 1
        resp, err = _do_call(user if attempts == 1 else user + _retry_suffix(validation_msg))
        if err:
            last_err = err
            break
        raw_text = _extract_text(resp)
        last_raw = raw_text
        parsed = extract_first_json_object(raw_text)
        if parsed is None:
            validation_msg = "JSON_PARSE_FAILED"
        else:
            if required_out_keys:
                ok, msg = _validate_parsed_judgements(parsed, required_out_keys)
                validation_msg = msg
                if ok:
                    break
            else:
                break

    meta: Dict[str, Any] = {
        "ok": parsed is not None and (validation_msg in ("", "OK")),
        "elapsed_sec": round(time.time() - t0, 3),
        "raw_len": len(last_raw),
        "model_used": model_used,
        "attempts": attempts,
        "validation": validation_msg,
    }

    if parsed is None:
        meta["error"] = last_err or validation_msg or "JSON_PARSE_FAILED"
        meta["raw_head"] = (last_raw or "")[:800]
    else:
        # if parsed exists but validation failed even after retries
        if required_out_keys and validation_msg not in ("", "OK"):
            meta["error"] = validation_msg
            meta["raw_head"] = (last_raw or "")[:800]

    return parsed, meta


def _retry_suffix(validation_msg: str) -> str:
    return (
        "\n\n【再出力指示】\n"
        "前の出力は要件を満たしていません。以下を必ず守って、同じJSON形式で出力し直してください。\n"
        "- JSONオブジェクトのみ（説明文禁止）\n"
        "- judgements 内の全項目で winner は A/B/tie のいずれか\n"
        "- confidence は 0.01〜1.0 の数値（0.0は禁止）\n"
        "- reason は必ず日本語で1〜3文。空欄禁止。\n"
        f"- 失敗理由: {validation_msg}\n"
    )


# ----------------------------
# Aggregation helpers
# ----------------------------
def expected_target_cond_for_key(key: str, A_cond: str, B_cond: str, cond_specs: Dict[str, CondSpec]) -> str:
    k = str(key or "").strip()

    def pick(pred) -> str:
        a = cond_specs.get(A_cond)
        b = cond_specs.get(B_cond)
        if a and pred(a):
            return A_cond
        if b and pred(b):
            return B_cond
        return ""

    if k == "summary":
        return pick(lambda c: c.detail == "summary")
    if k == "detail":
        return pick(lambda c: c.detail == "detail")
    if k == "intro":
        return pick(lambda c: c.level == "intro")
    if k == "advanced":
        return pick(lambda c: c.level == "advanced")
    if k == "modality_audio_fitness":
        return pick(lambda c: c.type == "audio")
    # goal系は「どちらが良いか」なので、狙い（通常 combo 側）を target として置く
    if k == "goal_deep_listening":
        return "combo_adv_detail_anim" if "combo_adv_detail_anim" in {A_cond, B_cond} else ""
    if k == "goal_overview":
        return "combo_intro_summary_audio" if "combo_intro_summary_audio" in {A_cond, B_cond} else ""
    if k == "goal_quick":
        return "combo_intro_summary_audio" if "combo_intro_summary_audio" in {A_cond, B_cond} else ""

    return ""


def aggregate_rep_payloads(rep_payloads: List[Dict[str, Any]], A_cond: str, B_cond: str, cond_specs: Dict[str, CondSpec]) -> Dict[str, Any]:
    """
    Aggregate judgements across repeats.
    Returns dict with per-key counts/mean/std and target_win_rate.
    """
    per_key: Dict[str, Dict[str, Any]] = {}
    for payload in rep_payloads:
        judgements = payload.get("judgements")
        if not isinstance(judgements, dict):
            continue
        for key, j in judgements.items():
            if not isinstance(j, dict):
                continue
            k = str(key)
            per_key.setdefault(k, {"winner": [], "confidence": []})
            per_key[k]["winner"].append(str(j.get("winner", "")).strip())
            per_key[k]["confidence"].append(safe_float(j.get("confidence", 0.0), 0.0))

    out: Dict[str, Any] = {"keys": {}, "A": A_cond, "B": B_cond}
    for k, d in per_key.items():
        winners = [w for w in d.get("winner", []) if w]
        confs = [c for c in d.get("confidence", []) if isinstance(c, (int, float))]
        cnt = Counter(winners)
        m, s = mean_std(confs)
        # majority
        maj = ""
        if winners:
            top = cnt.most_common()
            if len(top) >= 2 and top[0][1] == top[1][1]:
                maj = "tie"
            else:
                maj = top[0][0]
        target = expected_target_cond_for_key(k, A_cond, B_cond, cond_specs)
        # target win rate
        def winner_ab_to_cond(w: str) -> str:
            if w == "A":
                return A_cond
            if w == "B":
                return B_cond
            if w == "tie":
                return "tie"
            return "unknown"

        valid = 0
        hit = 0
        for w in winners:
            wc = winner_ab_to_cond(w)
            if not target or wc in ("tie", "unknown"):
                continue
            valid += 1
            if wc == target:
                hit += 1

        out["keys"][k] = {
            "n": len(winners),
            "counts": {
                "A": cnt.get("A", 0),
                "B": cnt.get("B", 0),
                "tie": cnt.get("tie", 0),
                "other": sum(cnt.values()) - cnt.get("A", 0) - cnt.get("B", 0) - cnt.get("tie", 0),
            },
            "mean_confidence": m,
            "std_confidence": s,
            "majority_winner": maj,
            "target_cond": target,
            "target_win_rate": (hit / valid) if valid else "",
            "target_n": valid,
        }
    return out


# ----------------------------
# Excel export
# ----------------------------
def _ensure_openpyxl() -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not available. Please install openpyxl.")


def _ws_write_table(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    # header style
    bold = Font(bold=True) if Font else None
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        if bold:
            cell.font = bold
        cell.alignment = Alignment(vertical="top", wrap_text=True) if Alignment else None

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # width
    for i, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            v = row[0].value
            if v is None:
                continue
            max_len = max(max_len, min(80, len(str(v))))
        if get_column_letter:
            ws.column_dimensions[get_column_letter(i)].width = min(80, max(10, max_len + 2))

    # wrap long text-ish
    wrap_cols = set()
    for i, h in enumerate(headers, start=1):
        if any(k in h for k in ("reason", "notes", "raw_head")):
            wrap_cols.add(i)
    if wrap_cols and Alignment:
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in wrap_cols:
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)


def export_step6_excel(run_dir: Path, step6_csv: Path, excel_path: Path, cfg_path: Path) -> Path:
    _ensure_openpyxl()

    if not step6_csv.exists():
        raise FileNotFoundError(f"Step6 csv not found: {step6_csv}")
    cfg = read_json(cfg_path)
    cond_specs = {k: CondSpec(**v) for k, v in cfg.get("condition_specs", {}).items()}

    # read csv rows
    with step6_csv.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    # long rows: one row per (trial, key)
    long_rows: List[Dict[str, Any]] = []
    err_rows: List[Dict[str, Any]] = []

    for r in csv_rows:
        ok = str(r.get("ok", "0")).strip() == "1"
        jp = resolve_json_path_for_row(r, run_dir)
        if not str(jp).strip():
            if not ok:
                err_rows.append({
                    "run_id": r.get("run_id", ""),
                    "model": r.get("model", ""),
                    "lecture": r.get("lecture", ""),
                    "pair_id": r.get("pair_id", ""),
                    "rep": r.get("rep", ""),
                    "error": r.get("error", "") or "NO_JSON_PATH",
                    "timestamp": r.get("timestamp", ""),
                })
            continue

        if not jp.exists():
            err_rows.append({
                "run_id": r.get("run_id", ""),
                "model": r.get("model", ""),
                "lecture": r.get("lecture", ""),
                "pair_id": r.get("pair_id", ""),
                "rep": r.get("rep", ""),
                "error": "JSON_FILE_MISSING",
                "json_path": str(jp),
                "timestamp": r.get("timestamp", ""),
            })
            continue

        try:
            payload = read_json(jp)
        except Exception as e:
            err_rows.append({
                "run_id": r.get("run_id", ""),
                "model": r.get("model", ""),
                "lecture": r.get("lecture", ""),
                "pair_id": r.get("pair_id", ""),
                "rep": r.get("rep", ""),
                "error": f"JSON_READ_FAILED: {type(e).__name__}: {e}",
                "json_path": str(jp),
                "timestamp": r.get("timestamp", ""),
            })
            continue

        A_cond = str(payload.get("A", "") or r.get("A_cond", ""))
        B_cond = str(payload.get("B", "") or r.get("B_cond", ""))

        judgements = payload.get("judgements")
        if not isinstance(judgements, dict):
            if not ok:
                err_rows.append({
                    "run_id": r.get("run_id", ""),
                    "model": r.get("model", ""),
                    "lecture": r.get("lecture", ""),
                    "pair_id": r.get("pair_id", ""),
                    "rep": r.get("rep", ""),
                    "error": r.get("error", "") or "JUDGEMENTS_MISSING",
                    "json_path": str(jp),
                    "timestamp": r.get("timestamp", ""),
                })
            continue

        for key, j in judgements.items():
            if not isinstance(j, dict):
                continue
            winner_raw = str(j.get("winner", "")).strip()
            conf = safe_float(j.get("confidence", 0.0), 0.0)
            reason = str(j.get("reason", ""))

            # winner cond
            if winner_raw == "A":
                winner_cond = A_cond
            elif winner_raw == "B":
                winner_cond = B_cond
            elif winner_raw == "tie":
                winner_cond = "tie"
            else:
                winner_cond = "unknown"

            target_cond = expected_target_cond_for_key(str(key), A_cond, B_cond, cond_specs)
            is_target = ""
            if target_cond and winner_cond not in ("tie", "unknown"):
                is_target = 1 if winner_cond == target_cond else 0

            long_rows.append({
                "run_id": r.get("run_id", ""),
                "model": r.get("model", ""),
                "lecture": r.get("lecture", ""),
                "pair_id": r.get("pair_id", ""),
                "rep": r.get("rep", ""),
                "key": str(key),
                "A_cond": A_cond,
                "B_cond": B_cond,
                "winner_raw": winner_raw,
                "winner_cond": winner_cond,
                "confidence": conf,
                "reason": reason,
                "target_cond": target_cond,
                "is_target": is_target,
                "json_path": str(jp),
                "timestamp": r.get("timestamp", ""),
            })

    # --- aggregate: pair × key (averaging over repeats)
    agg_pair_key_rows: List[Dict[str, Any]] = []
    by_pk: Dict[Tuple[str, str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for lr in long_rows:
        by_pk[(lr["model"], lr["lecture"], lr["pair_id"], lr["key"])].append(lr)

    for (model, lecture, pair_id, key), items in sorted(by_pk.items(), key=lambda x: (x[0][0], x[0][1], x[0][2], x[0][3])):
        winners = [str(it.get("winner_raw", "")).strip() for it in items if str(it.get("winner_raw", "")).strip()]
        confs = [safe_float(it.get("confidence", 0.0), 0.0) for it in items]
        m, s = mean_std(confs)
        cnt = Counter(winners)
        # majority
        maj = ""
        if winners:
            top = cnt.most_common()
            if len(top) >= 2 and top[0][1] == top[1][1]:
                maj = "tie"
            else:
                maj = top[0][0]
        # target rate
        is_targets = [it.get("is_target") for it in items if it.get("is_target") in (0, 1)]
        twr = (sum(is_targets) / len(is_targets)) if is_targets else ""

        # carry A/B/target
        A_cond = items[0].get("A_cond", "")
        B_cond = items[0].get("B_cond", "")
        target_cond = items[0].get("target_cond", "")

        agg_pair_key_rows.append({
            "model": model,
            "lecture": lecture,
            "pair_id": pair_id,
            "key": key,
            "n_trials": len(items),
            "A": cnt.get("A", 0),
            "B": cnt.get("B", 0),
            "tie": cnt.get("tie", 0),
            "mean_confidence": m,
            "std_confidence": s,
            "majority_winner": maj,
            "A_cond": A_cond,
            "B_cond": B_cond,
            "target_cond": target_cond,
            "target_win_rate": twr,
        })

    # --- aggregate: model × key (user request)
    agg_model_key_rows: List[Dict[str, Any]] = []
    by_mk: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for lr in long_rows:
        by_mk[(lr["model"], lr["key"])].append(lr)

    for (model, key), items in sorted(by_mk.items(), key=lambda x: (x[0][0], x[0][1])):
        winners = [str(it.get("winner_raw", "")).strip() for it in items if str(it.get("winner_raw", "")).strip()]
        confs = [safe_float(it.get("confidence", 0.0), 0.0) for it in items]
        m, s = mean_std(confs)
        cnt = Counter(winners)
        is_targets = [it.get("is_target") for it in items if it.get("is_target") in (0, 1)]
        twr = (sum(is_targets) / len(is_targets)) if is_targets else ""
        agg_model_key_rows.append({
            "model": model,
            "key": key,
            "n_trials": len(items),
            "A": cnt.get("A", 0),
            "B": cnt.get("B", 0),
            "tie": cnt.get("tie", 0),
            "mean_confidence": m,
            "std_confidence": s,
            "target_win_rate": twr,
        })

    # --- write workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_info = wb.create_sheet("info")
    _ws_write_table(ws_info, ["key", "value"], [
        {"key": "run_dir", "value": str(run_dir)},
        {"key": "config_used", "value": str(cfg_path)},
        {"key": "step6_csv", "value": str(step6_csv)},
        {"key": "generated_at", "value": now_iso()},
        {"key": "n_csv_rows", "value": len(csv_rows)},
        {"key": "n_long_rows", "value": len(long_rows)},
    ])

    ws_raw = wb.create_sheet("raw_csv")
    raw_headers = list(csv_rows[0].keys()) if csv_rows else []
    _ws_write_table(ws_raw, raw_headers, csv_rows)

    ws_long = wb.create_sheet("judgements_long")
    long_headers = [
        "run_id", "model", "lecture", "pair_id", "rep",
        "key", "A_cond", "B_cond",
        "winner_raw", "winner_cond",
        "confidence", "reason",
        "target_cond", "is_target",
        "json_relpath", "json_path", "timestamp",
    ]
    _ws_write_table(ws_long, long_headers, long_rows)

    ws_pair = wb.create_sheet("summary_pair_key")
    pair_headers = [
        "model", "lecture", "pair_id", "key",
        "n_trials", "A", "B", "tie",
        "mean_confidence", "std_confidence", "majority_winner",
        "A_cond", "B_cond", "target_cond", "target_win_rate",
    ]
    _ws_write_table(ws_pair, pair_headers, agg_pair_key_rows)

    ws_mk = wb.create_sheet("summary_model_key")
    mk_headers = [
        "model", "key", "n_trials",
        "A", "B", "tie",
        "mean_confidence", "std_confidence", "target_win_rate",
    ]
    _ws_write_table(ws_mk, mk_headers, agg_model_key_rows)

    ws_err = wb.create_sheet("errors")
    err_headers = ["run_id", "model", "lecture", "pair_id", "rep", "error", "json_path", "timestamp"]
    _ws_write_table(ws_err, err_headers, err_rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    return excel_path


# ----------------------------
# Main
# ----------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-id", required=True)

    ap.add_argument("--template-config", default="", help="override config path")
    ap.add_argument("--manifest", default="", help="override manifest_step02.jsonl path")
    ap.add_argument("--apikey", default="", help="override apikey.txt path (default: <project_root>/apikey.txt)")

    ap.add_argument("--allow-partial", action="store_true", help="continue even if missing scripts; record errors")
    ap.add_argument("--force", action="store_true", help="recreate json even if exists")
    ap.add_argument("--sleep", type=float, default=0.0)

    # repeat settings (user request: 10)
    ap.add_argument("--repeats", type=int, default=10, help="number of repeated calls per (lecture,pair,model)")

    # models: default fixed (GPT3.5/GPT4o/GPT4o-mini/GPT5.2). keep config list if needed.
    ap.add_argument("--use-config-models", action="store_true", help="use judge_models from config (do not override)")

    # filters for debugging
    ap.add_argument("--lecture", action="append", default=[], help="only this lecture (repeatable)")
    ap.add_argument("--pair", action="append", default=[], help="only this pair_id (repeatable)")
    ap.add_argument("--model", action="append", default=[], help="only these judge model IDs (repeatable)")

    # optional: append models
    ap.add_argument("--add-model", action="append", default=[], help="append judge model(s) in addition")

    # Excel
    ap.add_argument("--no-excel", action="store_true", help="do not export pair_judgements.xlsx")
    ap.add_argument("--excel-path", default="", help="override excel output path")
    ap.add_argument("--excel-only", action="store_true", help="only export excel from existing pair_judgements.csv/json")

    args = ap.parse_args()

    run_dir = RUNS_ROOT / args.run_id
    if not run_dir.exists():
        raise FileNotFoundError(f"run_dir not found: {run_dir}")

    out_csv = run_dir / "analysis" / "pair_judgements.csv"
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    excel_path = Path(args.excel_path) if args.excel_path.strip() else (run_dir / "analysis" / "pair_judgements.xlsx")
    if not excel_path.is_absolute():
        excel_path = (Path(".") / excel_path).resolve()

    cfg_path = resolve_config_path(args.template_config, run_dir)
    if not cfg_path.exists():
        raise FileNotFoundError(f"config not found: {cfg_path}")

    if args.excel_only:
        if args.no_excel:
            print("[Step6] --excel-only but --no-excel set. Nothing to do.")
            return
        if not out_csv.exists():
            raise FileNotFoundError(f"pair_judgements.csv not found: {out_csv}")
        p = export_step6_excel(run_dir, out_csv, excel_path, cfg_path)
        print(f"[Step6] Excel exported: {p}")
        return

    judge_models_cfg, pairs, cond_specs = load_template_config(cfg_path)

    # models
    if args.use_config_models:
        judge_models = [normalize_model_id(m) for m in judge_models_cfg]
    else:
        judge_models = [normalize_model_id(m) for m in DEFAULT_FIXED_MODELS]

    # append
    if args.add_model:
        judge_models = judge_models + [normalize_model_id(m) for m in args.add_model]

    # de-dup keep order
    seen = set()
    judge_models = [m for m in judge_models if (m not in seen and not seen.add(m))]

    # filter by args.model if given
    if args.model:
        allow = {normalize_model_id(m) for m in args.model}
        judge_models = [m for m in judge_models if m in allow]
    if not judge_models:
        raise RuntimeError("No judge_models after filtering.")

    # filter pairs
    if args.pair:
        allow_pairs = set(args.pair)
        pairs = [p for p in pairs if p.pair_id in allow_pairs]
    if not pairs:
        raise RuntimeError("No pairs after filtering.")

    manifest_path = Path(args.manifest) if args.manifest else (run_dir / "generation" / "manifest_step02.jsonl")
    if not manifest_path.is_absolute():
        manifest_path = (Path(".") / manifest_path).resolve()

    rows = load_manifest_jsonl(manifest_path)
    idx = build_index(rows, run_dir)

    allow_lectures = set(args.lecture) if args.lecture else None

    apikey_path: Optional[Path] = Path(args.apikey) if args.apikey.strip() else None
    client = create_client(apikey_path)

    fieldnames = [
        "run_id", "model",
        "lecture", "pair_id", "rep",
        "A_cond", "B_cond",
        "A_mode", "B_mode",
        "keys",
        "ok", "elapsed_sec", "raw_len", "error",
        "json_relpath", "json_path", "timestamp",
    ]

    write_header = not out_csv.exists()
    with out_csv.open("a", encoding="utf-8", newline="") as fcsv:
        writer = csv.DictWriter(fcsv, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()

        lectures_all = sorted({r.lecture for r in rows if r.lecture})
        lectures = [L for L in lectures_all if (allow_lectures is None or L in allow_lectures)]
        if not lectures:
            raise RuntimeError("No lectures to process (after filter).")

        repeats = max(1, int(args.repeats))
        total = len(lectures) * len(pairs) * len(judge_models) * repeats
        step = 0

        for lecture_key in lectures:
            for pair in pairs:
                A_cond = pair.A
                B_cond = pair.B

                if A_cond not in cond_specs or B_cond not in cond_specs:
                    # write a single error row (rep empty)
                    for model in judge_models:
                        writer.writerow({
                            "run_id": args.run_id,
                            "model": model,
                            "lecture": lecture_key,
                            "pair_id": pair.pair_id,
                            "rep": "",
                            "A_cond": A_cond,
                            "B_cond": B_cond,
                            "A_mode": "",
                            "B_mode": "",
                            "keys": "",
                            "ok": "0",
                            "elapsed_sec": "0",
                            "raw_len": "0",
                            "error": "COND_SPEC_MISSING",
                            "json_relpath": "",
                            "json_path": "",
                            "timestamp": now_iso(),
                        })
                    if not args.allow_partial:
                        raise RuntimeError(f"condition_specs missing for {A_cond} or {B_cond}")
                    continue

                keys = needed_judgements(pair.pair_id, A_cond, B_cond, cond_specs[A_cond], cond_specs[B_cond])
                if not keys:
                    continue

                A_root = idx.get((lecture_key, A_cond))
                B_root = idx.get((lecture_key, B_cond))

                if A_root is None or B_root is None:
                    for model in judge_models:
                        writer.writerow({
                            "run_id": args.run_id,
                            "model": model,
                            "lecture": lecture_key,
                            "pair_id": pair.pair_id,
                            "rep": "",
                            "A_cond": A_cond,
                            "B_cond": B_cond,
                            "A_mode": "",
                            "B_mode": "",
                            "keys": ",".join(keys),
                            "ok": "0",
                            "elapsed_sec": "0",
                            "raw_len": "0",
                            "error": "DEST_NOT_FOUND_IN_MANIFEST",
                            "json_relpath": "",
                            "json_path": "",
                            "timestamp": now_iso(),
                        })
                    if not args.allow_partial:
                        raise RuntimeError(f"dest not found in manifest for {lecture_key} {A_cond}/{B_cond}")
                    continue

                A_mode = detect_mode(A_root)
                B_mode = detect_mode(B_root)

                A_script, A_src = pick_script_path(A_root, A_mode)
                B_script, B_src = pick_script_path(B_root, B_mode)

                if A_script is None or B_script is None or (not A_script.exists()) or (not B_script.exists()):
                    for model in judge_models:
                        writer.writerow({
                            "run_id": args.run_id,
                            "model": model,
                            "lecture": lecture_key,
                            "pair_id": pair.pair_id,
                            "rep": "",
                            "A_cond": A_cond,
                            "B_cond": B_cond,
                            "A_mode": A_mode,
                            "B_mode": B_mode,
                            "keys": ",".join(keys),
                            "ok": "0",
                            "elapsed_sec": "0",
                            "raw_len": "0",
                            "error": "SCRIPT_NOT_FOUND",
                            "json_relpath": "",
                            "json_path": "",
                            "timestamp": now_iso(),
                        })
                    if not args.allow_partial:
                        raise FileNotFoundError(f"script not found for pair {pair.pair_id} lecture {lecture_key}")
                    continue

                A_text = read_text_best_effort(A_script)
                B_text = read_text_best_effort(B_script)

                system, user = build_pair_prompt(
                    lecture_key=lecture_key,
                    pair_id=pair.pair_id,
                    A_cond=A_cond,
                    B_cond=B_cond,
                    A_text=A_text,
                    B_text=B_text,
                    keys=keys,
                )

                for model in judge_models:
                    safe_lecture = sanitize_name(lecture_key)
                    safe_pair = sanitize_name(pair.pair_id)
                    out_dir = run_dir / "analysis" / "pair_judgements" / model / safe_lecture / safe_pair
                    out_dir.mkdir(parents=True, exist_ok=True)

                    rep_payloads: List[Dict[str, Any]] = []

                    for rep in range(1, repeats + 1):
                        step += 1
                        print(
                            f"[Step6] ({step}/{total}) lecture={lecture_key} pair={pair.pair_id} model={model} rep={rep}/{repeats} keys={keys}",
                            flush=True,
                        )

                        out_json = out_dir / f"rep_{rep:02d}.json"

                        if out_json.exists() and not args.force:
                            writer.writerow({
                                "run_id": args.run_id,
                                "model": model,
                                "lecture": lecture_key,
                                "pair_id": pair.pair_id,
                                "rep": str(rep),
                                "A_cond": A_cond,
                                "B_cond": B_cond,
                                "A_mode": A_mode,
                                "B_mode": B_mode,
                                "keys": ",".join(keys),
                                "ok": "1",
                                "elapsed_sec": "0",
                                "raw_len": "0",
                                "error": "CACHE_HIT",
                                "json_relpath": str(_relpath_from_run(out_json, run_dir)),
                                "json_path": str(out_json),
                                "timestamp": now_iso(),
                            })
                            try:
                                rep_payloads.append(read_json(out_json))
                            except Exception:
                                pass
                            continue

                        required_out_keys = expand_out_keys(keys)
                        parsed, meta = call_llm(
                            client,
                            model,
                            system,
                            user,
                            required_out_keys=required_out_keys,
                            max_retries=1,
                        )

                        payload: Dict[str, Any] = {
                            "lecture": lecture_key,
                            "pair_id": pair.pair_id,
                            "A": A_cond,
                            "B": B_cond,
                            "keys": keys,
                            "judgements": None if parsed is None else parsed.get("judgements", None),
                            "notes": "" if parsed is None else str(parsed.get("notes", "")),
                            "_meta": {
                                "run_id": args.run_id,
                                "model": model,
                                "rep": rep,
                                "A_mode": A_mode,
                                "B_mode": B_mode,
                                "A_script_source": A_src,
                                "B_script_source": B_src,
                                "A_script_path": str(A_script),
                                "B_script_path": str(B_script),
                                "llm": meta,
                                "timestamp": now_iso(),
                            },
                        }

                        if parsed is None:
                            payload["_meta"]["error"] = meta.get("error", "UNKNOWN_ERROR")
                            if "raw_head" in meta:
                                payload["_meta"]["raw_head"] = meta["raw_head"]
                        else:
                            payload["_meta"]["parsed_raw"] = parsed

                        write_json(out_json, payload)
                        rep_payloads.append(payload)

                        ok = parsed is not None
                        writer.writerow({
                            "run_id": args.run_id,
                            "model": model,
                            "lecture": lecture_key,
                            "pair_id": pair.pair_id,
                            "rep": str(rep),
                            "A_cond": A_cond,
                            "B_cond": B_cond,
                            "A_mode": A_mode,
                            "B_mode": B_mode,
                            "keys": ",".join(keys),
                            "ok": "1" if ok else "0",
                            "elapsed_sec": str(meta.get("elapsed_sec", 0)),
                            "raw_len": str(meta.get("raw_len", 0)),
                            "error": "" if ok else str(meta.get("error", "UNKNOWN_ERROR")),
                            "json_relpath": str(_relpath_from_run(out_json, run_dir)),
                            "json_path": str(out_json),
                            "timestamp": now_iso(),
                        })

                        if args.sleep and args.sleep > 0:
                            time.sleep(args.sleep)

                    # write aggregate.json (per model/pair/lecture)
                    try:
                        agg = aggregate_rep_payloads(rep_payloads, A_cond, B_cond, cond_specs)
                        agg["_meta"] = {
                            "run_id": args.run_id,
                            "model": model,
                            "lecture": lecture_key,
                            "pair_id": pair.pair_id,
                            "repeats": repeats,
                            "generated_at": now_iso(),
                        }
                        write_json(out_dir / "aggregate.json", agg)
                    except Exception:
                        pass

    # export Excel
    if not args.no_excel:
        try:
            p = export_step6_excel(run_dir, out_csv, excel_path, cfg_path)
            print(f"[Step6] Excel exported: {p}")
        except Exception as e:
            print(f"[Step6] Excel export failed: {type(e).__name__}: {e}")

    print("[Step6] done.")


if __name__ == "__main__":
    main()
