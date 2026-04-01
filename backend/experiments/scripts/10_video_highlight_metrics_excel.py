# experiments/scripts/10_video_highlight_metrics_excel.py
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ============================================================
# Paths
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parents[2]  # auto_lecture/
RUNS_ROOT = PROJECT_ROOT / "experiments" / "runs"


# ============================================================
# Small utils
# ============================================================
def log(msg: str) -> None:
    print(f"[StepX][video_hl_metrics] {msg}", flush=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_text_best_effort(path: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp932"):
        try:
            return path.read_text(encoding=enc)
        except Exception:
            continue
    return path.read_text(errors="ignore")


def safe_int(x: Any, default: int = 0) -> int:
    try:
        return int(str(x).strip())
    except Exception:
        return default


def safe_float(x: Any, default: float = 0.0) -> float:
    try:
        return float(str(x).strip())
    except Exception:
        return default


# ============================================================
# Media duration (ffprobe -> opencv fallback for mp4)
# ============================================================
def _ffprobe_duration_sec(path: Path) -> Optional[float]:
    try:
        cmd = [
            "ffprobe", "-v", "error",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1",
            str(path),
        ]
        out = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True).strip()
        if out:
            v = float(out)
            if v > 0:
                return v
    except Exception:
        return None
    return None


def _opencv_mp4_duration_sec(path: Path) -> Optional[float]:
    try:
        import cv2  # type: ignore
        cap = cv2.VideoCapture(str(path))
        if not cap.isOpened():
            return None
        fps = cap.get(cv2.CAP_PROP_FPS)
        n = cap.get(cv2.CAP_PROP_FRAME_COUNT)
        cap.release()
        if fps and fps > 0 and n and n > 0:
            return float(n) / float(fps)
    except Exception:
        return None
    return None


def _mutagen_audio_duration_sec(path: Path) -> Optional[float]:
    # ffprobe が無い環境向け保険（入っていれば使う）
    try:
        from mutagen import File  # type: ignore
        a = File(str(path))
        if a is not None and a.info is not None and getattr(a.info, "length", None):
            v = float(a.info.length)
            if v > 0:
                return v
    except Exception:
        return None
    return None


def get_media_duration_sec(path: Path) -> Optional[float]:
    if not path.exists():
        return None
    d = _ffprobe_duration_sec(path)
    if d is not None:
        return d
    # mp4 のみ OpenCV fallback
    if path.suffix.lower() == ".mp4":
        d = _opencv_mp4_duration_sec(path)
        if d is not None:
            return d
    # 音声は mutagen fallback
    if path.suffix.lower() in (".mp3", ".m4a", ".wav"):
        d = _mutagen_audio_duration_sec(path)
        if d is not None:
            return d
    return None


def hms_str(sec: Optional[float]) -> str:
    if sec is None or not math.isfinite(sec) or sec < 0:
        return ""
    s = int(round(sec))
    hh = s // 3600
    mm = (s % 3600) // 60
    ss = s % 60
    return f"{hh:02d}:{mm:02d}:{ss:02d}"


# ============================================================
# Detect files in output_root
# ============================================================
RE_ANIM_MP4 = re.compile(r"^slide_(\d{3})_sent(\d{2})(?:_(.+))?\.mp4$", re.IGNORECASE)


def find_final_video(output_root: Path) -> Optional[Path]:
    out_final = output_root / "lecture_outputs" / "output_final"
    if not out_final.exists():
        return None

    # よくありそうな名前を優先
    for name in ("lecture_final.mp4", "final.mp4", "output.mp4", "lecture.mp4"):
        p = out_final / name
        if p.exists():
            return p

    # 無ければ一番大きい mp4
    mp4s = [p for p in out_final.rglob("*.mp4") if p.is_file()]
    if not mp4s:
        return None
    mp4s.sort(key=lambda p: p.stat().st_size, reverse=True)
    return mp4s[0]


def find_audio_duration(output_root: Path) -> Tuple[Optional[float], str]:
    # 1) output_final に単体音声があればそれ優先
    out_final = output_root / "lecture_outputs" / "output_final"
    if out_final.exists():
        auds = [p for p in out_final.rglob("*") if p.suffix.lower() in (".mp3", ".m4a", ".wav")]
        if auds:
            auds.sort(key=lambda p: p.stat().st_size, reverse=True)
            d = get_media_duration_sec(auds[0])
            return d, f"output_final:{auds[0].name}"

    # 2) tts_outputs の mp3 群を合算（page/part 分割を想定）
    tts_dir = output_root / "lecture_outputs" / "tts_outputs"
    if tts_dir.exists():
        mp3s = [p for p in tts_dir.rglob("*.mp3") if p.is_file()]
        if mp3s:
            ds = []
            miss = 0
            for p in mp3s:
                d = get_media_duration_sec(p)
                if d is None:
                    miss += 1
                else:
                    ds.append(d)
            if ds:
                return float(sum(ds)), f"tts_outputs:sum_mp3(n={len(mp3s)},missing={miss})"

    return None, "not_found"


@dataclass(frozen=True)
class IndexRow:
    lecture_title: str
    cond_id: str
    status: str
    meta_json: str


def load_extracted_index(index_csv: Path) -> List[IndexRow]:
    rows: List[IndexRow] = []
    with index_csv.open("r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for d in r:
            rows.append(
                IndexRow(
                    lecture_title=str(d.get("lecture_title") or ""),
                    cond_id=str(d.get("cond_id") or ""),
                    status=str(d.get("status") or ""),
                    meta_json=str(d.get("meta_json") or ""),
                )
            )
    return rows


def load_run_condition_specs(run_dir: Path) -> Dict[str, Dict[str, Any]]:
    cfg = run_dir / "config" / "experiment_config.json"
    if not cfg.exists():
        return {}
    obj = read_json(cfg)
    cond_specs = obj.get("condition_specs", {})
    return cond_specs if isinstance(cond_specs, dict) else {}


# ============================================================
# Binning
# ============================================================
def add_time_bins(intervals: List[Tuple[float, float]], total_sec: float, bin_sec: int) -> pd.DataFrame:
    if total_sec <= 0:
        return pd.DataFrame(columns=["bin_start_sec", "bin_end_sec", "highlight_sec"])
    nb = int(math.ceil(total_sec / bin_sec))
    bins = [(i * bin_sec, min((i + 1) * bin_sec, total_sec)) for i in range(nb)]
    acc = [0.0] * nb
    for s, e in intervals:
        s = max(0.0, min(s, total_sec))
        e = max(0.0, min(e, total_sec))
        if e <= s:
            continue
        b0 = int(s // bin_sec)
        b1 = int((e - 1e-9) // bin_sec)
        b0 = max(0, min(b0, nb - 1))
        b1 = max(0, min(b1, nb - 1))
        for bi in range(b0, b1 + 1):
            bs, be = bins[bi]
            ov = max(0.0, min(e, be) - max(s, bs))
            if ov > 0:
                acc[bi] += ov
    return pd.DataFrame(
        [{"bin_start_sec": bs, "bin_end_sec": be, "highlight_sec": acc[i]} for i, (bs, be) in enumerate(bins)]
    )


def add_percent_bins(intervals: List[Tuple[float, float]], total_sec: float, n_bins: int) -> pd.DataFrame:
    if total_sec <= 0:
        return pd.DataFrame(columns=["bin_idx", "bin_start_pct", "bin_end_pct", "highlight_sec"])
    edges = [i / n_bins for i in range(n_bins + 1)]
    acc = [0.0] * n_bins
    for s, e in intervals:
        s = max(0.0, min(s, total_sec)) / total_sec
        e = max(0.0, min(e, total_sec)) / total_sec
        if e <= s:
            continue
        b0 = int(min(n_bins - 1, max(0, math.floor(s * n_bins))))
        b1 = int(min(n_bins - 1, max(0, math.floor((e - 1e-12) * n_bins))))
        for bi in range(b0, b1 + 1):
            bs, be = edges[bi], edges[bi + 1]
            ov = max(0.0, min(e, be) - max(s, bs))
            if ov > 0:
                acc[bi] += ov * total_sec
    return pd.DataFrame(
        [
            {
                "bin_idx": i,
                "bin_start_pct": edges[i] * 100.0,
                "bin_end_pct": edges[i + 1] * 100.0,
                "highlight_sec": acc[i],
            }
            for i in range(n_bins)
        ]
    )


# ============================================================
# Excel helpers
# ============================================================
def autosize_columns(ws, max_width: int = 45) -> None:
    for col in ws.columns:
        values = []
        for cell in col[:200]:  # 上位だけで十分
            v = "" if cell.value is None else str(cell.value)
            values.append(len(v))
        width = min(max_width, max(10, max(values) + 2 if values else 10))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def add_heatmap(ws, start_cell: str, end_cell: str) -> None:
    rule = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="D9E1F2",
        end_type="max", end_color="2F5597",
    )
    ws.conditional_formatting.add(f"{start_cell}:{end_cell}", rule)


# ============================================================
# Core
# ============================================================
def build(run_id: str, runs_root: Path, bin_sec: int, pct_bins: int, allow_partial: bool, force: bool) -> Path:
    run_dir = runs_root / run_id
    if not run_dir.exists():
        raise FileNotFoundError(f"run_dir not found: {run_dir}")

    index_csv = run_dir / "extracted" / "index.csv"
    if not index_csv.exists():
        raise FileNotFoundError(f"extracted/index.csv not found: {index_csv} (run Step3 first)")

    cond_specs = load_run_condition_specs(run_dir)

    out_dir = run_dir / "analysis" / "video_highlight_metrics"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / "video_highlight_metrics.xlsx"
    if out_xlsx.exists() and force:
        out_xlsx.unlink(missing_ok=True)

    idx_rows = load_extracted_index(index_csv)

    summary_rows: List[Dict[str, Any]] = []
    raw_clip_rows: List[Dict[str, Any]] = []
    bin60_rows: List[Dict[str, Any]] = []
    binpct_rows: List[Dict[str, Any]] = []
    style_rows: List[Dict[str, Any]] = []
    bbox_rows: List[Dict[str, Any]] = []
    slide_rows: List[Dict[str, Any]] = []

    for r in idx_rows:
        if r.status != "ok":
            continue

        lecture = r.lecture_title
        cond_id = r.cond_id

        spec = cond_specs.get(cond_id, {})
        spec_type = str(spec.get("type") or "")
        spec_level = str(spec.get("level") or "")
        spec_detail = str(spec.get("detail") or "")

        meta_path = run_dir / Path(r.meta_json) if r.meta_json else None
        if meta_path is None or not meta_path.exists():
            if allow_partial:
                summary_rows.append({
                    "lecture": lecture, "cond_id": cond_id, "mode": "",
                    "duration_sec": "", "duration_hms": "",
                    "duration_source": "meta_missing",
                    "highlight_sec": "", "highlight_coverage": "",
                    "highlight_count": "", "highlight_rate": "",
                    "note": f"meta_json missing: {r.meta_json}",
                    "spec_type": spec_type, "spec_level": spec_level, "spec_detail": spec_detail,
                })
                continue
            raise FileNotFoundError(f"meta_json not found: {meta_path}")

        meta = read_json(meta_path)
        mode = str(meta.get("mode") or "")

        resolved_dest = str(meta.get("resolved_dest") or "").strip()
        output_root = Path(resolved_dest) if resolved_dest else None
        if output_root is None or not output_root.exists():
            if allow_partial:
                summary_rows.append({
                    "lecture": lecture, "cond_id": cond_id, "mode": mode,
                    "duration_sec": "", "duration_hms": "",
                    "duration_source": "dest_missing",
                    "highlight_sec": "", "highlight_coverage": "",
                    "highlight_count": "", "highlight_rate": "",
                    "note": f"resolved_dest missing: {resolved_dest}",
                    "spec_type": spec_type, "spec_level": spec_level, "spec_detail": spec_detail,
                })
                continue
            raise FileNotFoundError(f"resolved_dest not found: {resolved_dest}")

        # duration (prefer final mp4 if exists)
        final_mp4 = find_final_video(output_root)
        final_dur = get_media_duration_sec(final_mp4) if final_mp4 else None

        # --------------------------------------------
        # animation: parse per-sentence mp4 clips
        # --------------------------------------------
        if mode == "animation":
            anim_dir = output_root / "lecture_outputs" / "add_animation_outputs"
            mp4s = [p for p in anim_dir.glob("slide_*_sent*.mp4")] if anim_dir.exists() else []
            mp4_infos: List[Dict[str, Any]] = []
            miss_dur = 0

            for p in sorted(mp4s):
                m = RE_ANIM_MP4.match(p.name)
                if not m:
                    continue
                slide_no = int(m.group(1))
                sent_no = int(m.group(2))  # 1-based
                style = (m.group(3) or "").strip()
                d = get_media_duration_sec(p)
                if d is None:
                    miss_dur += 1
                mp4_infos.append({
                    "slide_no": slide_no,
                    "sentence_no": sent_no,
                    "anim_type_from_name": style,
                    "clip_path": str(p),
                    "clip_duration_sec": d,
                })

            df_mp4 = pd.DataFrame(mp4_infos)
            if df_mp4.empty:
                # ハイライト情報が無い（runner失敗など）
                dur_used = final_dur
                summary_rows.append({
                    "lecture": lecture, "cond_id": cond_id, "mode": mode,
                    "duration_sec": dur_used if dur_used is not None else "",
                    "duration_hms": hms_str(dur_used),
                    "duration_source": f"final_mp4:{final_mp4.name}" if final_mp4 else "unknown",
                    "highlight_sec": 0.0,
                    "highlight_coverage": 0.0,
                    "highlight_count": 0,
                    "highlight_rate": 0.0,
                    "highlight_center_pct": "",
                    "highlight_first_half_ratio": "",
                    "n_clips": 0,
                    "n_clips_missing_duration": miss_dur,
                    "note": "no add_animation_outputs mp4 found",
                    "spec_type": spec_type, "spec_level": spec_level, "spec_detail": spec_detail,
                })
                continue

            df_mp4 = df_mp4.sort_values(["slide_no", "sentence_no"]).reset_index(drop=True)

            # join extracted script_animation_table.csv (bbox_no, anim_type, text)
            tbl_path = run_dir / "extracted" / lecture / cond_id / "script_animation_table.csv"
            if tbl_path.exists():
                df_tbl = pd.read_csv(tbl_path)
            else:
                df_tbl = pd.DataFrame(columns=["slide_no", "sentence_no", "bbox_no", "anim_type", "text"])

            if not df_tbl.empty:
                # normalize dtypes
                df_tbl["slide_no"] = df_tbl["slide_no"].apply(lambda x: safe_int(x, 0))
                df_tbl["sentence_no"] = df_tbl["sentence_no"].apply(lambda x: safe_int(x, 0))

            df = df_mp4.merge(df_tbl, on=["slide_no", "sentence_no"], how="left")

            # choose anim_type (table preferred; fallback to filename)
            df["anim_type"] = df["anim_type"].fillna("").astype(str)
            df["bbox_no"] = df.get("bbox_no", "").fillna("").astype(str)
            df["text"] = df.get("text", "").fillna("").astype(str)
            df["anim_type"] = df["anim_type"].where(df["anim_type"].str.strip() != "", df["anim_type_from_name"].fillna("").astype(str))
            df["is_highlight"] = df["anim_type"].astype(str).str.strip() != ""

            # durations: if missing, set 0 (but keep count)
            df["clip_duration_sec"] = pd.to_numeric(df["clip_duration_sec"], errors="coerce")
            df["clip_duration_sec"] = df["clip_duration_sec"].fillna(0.0)

            # timeline
            df["start_sec"] = df["clip_duration_sec"].cumsum() - df["clip_duration_sec"]
            df["end_sec"] = df["clip_duration_sec"].cumsum()
            lecture_dur = float(df["clip_duration_sec"].sum())
            dur_source = "sum(add_animation_outputs)"
            if final_dur is not None and final_dur > 0:
                # 最終 mp4 の方が信頼できる場合は優先（ただし分布計算は clip timeline を使う）
                lecture_dur = float(final_dur)
                dur_source = f"final_mp4:{final_mp4.name}" if final_mp4 else "final_mp4"

            # highlight intervals (from clip timeline)
            hl = df[df["is_highlight"]].copy()
            hl_intervals = [(float(a), float(b)) for a, b in zip(hl["start_sec"], hl["end_sec"]) if b > a]
            hl_sec = float((hl["clip_duration_sec"]).sum())
            hl_count = int(hl.shape[0])

            coverage = (hl_sec / lecture_dur) if lecture_dur > 0 else 0.0
            hl_rate = (hl_count / (lecture_dur / 60.0)) if lecture_dur > 0 else 0.0

            # distribution stats
            center_pct = ""
            first_half_ratio = ""
            if lecture_dur > 0 and hl_sec > 0:
                mids = (hl["start_sec"] + hl["end_sec"]) / 2.0
                # weighted by clip_duration
                center = float((mids * hl["clip_duration_sec"]).sum() / hl_sec)
                center_pct = round(center / lecture_dur * 100.0, 2)
                first_half = float(hl[hl["end_sec"] <= (lecture_dur / 2.0)]["clip_duration_sec"].sum())
                first_half_ratio = round(first_half / hl_sec, 4)

            # raw clips
            for rec in df.to_dict(orient="records"):
                raw_clip_rows.append({
                    "run_id": run_id,
                    "lecture": lecture,
                    "cond_id": cond_id,
                    "mode": mode,
                    "slide_no": rec.get("slide_no", ""),
                    "sentence_no": rec.get("sentence_no", ""),
                    "anim_type": rec.get("anim_type", ""),
                    "bbox_no": rec.get("bbox_no", ""),
                    "is_highlight": int(bool(rec.get("is_highlight", False))),
                    "start_sec": round(float(rec.get("start_sec", 0.0)), 3),
                    "end_sec": round(float(rec.get("end_sec", 0.0)), 3),
                    "clip_duration_sec": round(float(rec.get("clip_duration_sec", 0.0)), 3),
                    "clip_path": rec.get("clip_path", ""),
                    "text": rec.get("text", ""),
                })

            # time bins (absolute)
            df_bin = add_time_bins(hl_intervals, lecture_dur, bin_sec)
            df_bin["run_id"] = run_id
            df_bin["lecture"] = lecture
            df_bin["cond_id"] = cond_id
            df_bin["mode"] = mode
            df_bin["bin_label"] = df_bin["bin_start_sec"].apply(lambda x: f"{int(x//60):02d}:{int(x%60):02d}")
            bin60_rows.extend(df_bin.to_dict(orient="records"))

            # percent bins
            df_pct = add_percent_bins(hl_intervals, lecture_dur, pct_bins)
            df_pct["run_id"] = run_id
            df_pct["lecture"] = lecture
            df_pct["cond_id"] = cond_id
            df_pct["mode"] = mode
            df_pct["bin_label"] = df_pct.apply(lambda r2: f"{int(r2['bin_start_pct']):02d}-{int(r2['bin_end_pct']):02d}%", axis=1)
            binpct_rows.extend(df_pct.to_dict(orient="records"))

            # style dist
            if hl_count > 0:
                g = hl.groupby("anim_type", dropna=False)["clip_duration_sec"].agg(["count", "sum"]).reset_index()
                for rec in g.to_dict(orient="records"):
                    style_rows.append({
                        "run_id": run_id, "lecture": lecture, "cond_id": cond_id,
                        "anim_type": rec.get("anim_type", ""),
                        "highlight_count": int(rec.get("count", 0)),
                        "highlight_sec": float(rec.get("sum", 0.0)),
                    })

            # bbox dist
            if "bbox_no" in hl.columns:
                hl2 = hl[hl["bbox_no"].astype(str).str.strip() != ""]
                if not hl2.empty:
                    g = hl2.groupby("bbox_no", dropna=False)["clip_duration_sec"].agg(["count", "sum"]).reset_index()
                    for rec in g.to_dict(orient="records"):
                        bbox_rows.append({
                            "run_id": run_id, "lecture": lecture, "cond_id": cond_id,
                            "bbox_no": rec.get("bbox_no", ""),
                            "highlight_count": int(rec.get("count", 0)),
                            "highlight_sec": float(rec.get("sum", 0.0)),
                        })

            # slide dist
            g = hl.groupby("slide_no", dropna=False)["clip_duration_sec"].agg(["count", "sum"]).reset_index()
            for rec in g.to_dict(orient="records"):
                slide_rows.append({
                    "run_id": run_id, "lecture": lecture, "cond_id": cond_id,
                    "slide_no": int(rec.get("slide_no", 0)),
                    "highlight_count": int(rec.get("count", 0)),
                    "highlight_sec": float(rec.get("sum", 0.0)),
                })

            summary_rows.append({
                "lecture": lecture,
                "cond_id": cond_id,
                "mode": mode,
                "spec_type": spec_type, "spec_level": spec_level, "spec_detail": spec_detail,
                "duration_sec": round(lecture_dur, 3) if lecture_dur > 0 else "",
                "duration_hms": hms_str(lecture_dur),
                "duration_source": dur_source,
                "n_clips": int(df.shape[0]),
                "n_highlight_clips": hl_count,
                "n_clips_missing_duration": miss_dur,
                "highlight_sec": round(hl_sec, 3),
                "highlight_coverage": round(coverage, 4),
                "highlight_count": hl_count,
                "highlight_rate": round(hl_rate, 4),  # per minute
                "highlight_center_pct": center_pct,
                "highlight_first_half_ratio": first_half_ratio,
                "final_mp4": str(final_mp4) if final_mp4 else "",
            })

        # --------------------------------------------
        # audio: duration only (highlight=0)
        # --------------------------------------------
        else:
            dur = None
            dur_source = "unknown"

            if final_dur is not None and final_dur > 0:
                dur = final_dur
                dur_source = f"final_mp4:{final_mp4.name}" if final_mp4 else "final_mp4"
            else:
                ad, src = find_audio_duration(output_root)
                dur = ad
                dur_source = src

            summary_rows.append({
                "lecture": lecture,
                "cond_id": cond_id,
                "mode": mode,
                "spec_type": spec_type, "spec_level": spec_level, "spec_detail": spec_detail,
                "duration_sec": round(dur, 3) if dur is not None else "",
                "duration_hms": hms_str(dur),
                "duration_source": dur_source,
                "n_clips": 0,
                "n_highlight_clips": 0,
                "n_clips_missing_duration": "",
                "highlight_sec": 0.0,
                "highlight_coverage": 0.0,
                "highlight_count": 0,
                "highlight_rate": 0.0,
                "highlight_center_pct": "",
                "highlight_first_half_ratio": "",
                "final_mp4": str(final_mp4) if final_mp4 else "",
            })

    # ============================================================
    # Build DataFrames
    # ============================================================
    df_summary = pd.DataFrame(summary_rows).sort_values(["lecture", "cond_id"], kind="stable")

    df_raw = pd.DataFrame(raw_clip_rows)
    if not df_raw.empty:
        df_raw = df_raw.sort_values(["lecture", "cond_id", "slide_no", "sentence_no"], kind="stable")

    df_bin60 = pd.DataFrame(bin60_rows)
    if not df_bin60.empty:
        df_bin60 = df_bin60.sort_values(["lecture", "cond_id", "bin_start_sec"], kind="stable")

    df_binpct = pd.DataFrame(binpct_rows)
    if not df_binpct.empty:
        df_binpct = df_binpct.sort_values(["lecture", "cond_id", "bin_idx"], kind="stable")

    df_style = pd.DataFrame(style_rows)
    if not df_style.empty:
        df_style = df_style.sort_values(["lecture", "cond_id", "highlight_sec"], ascending=[True, True, False])

    df_bbox = pd.DataFrame(bbox_rows)
    if not df_bbox.empty:
        df_bbox = df_bbox.sort_values(["lecture", "cond_id", "highlight_sec"], ascending=[True, True, False])

    df_slide = pd.DataFrame(slide_rows)
    if not df_slide.empty:
        df_slide = df_slide.sort_values(["lecture", "cond_id", "slide_no"], kind="stable")

    # heatmaps (pivot): highlight_sec
    df_heat60 = None
    if not df_bin60.empty:
        df_heat60 = df_bin60.pivot_table(
            index=["lecture", "cond_id"], columns="bin_label", values="highlight_sec", aggfunc="sum", fill_value=0.0
        ).reset_index()

    df_heatpct = None
    if not df_binpct.empty:
        df_heatpct = df_binpct.pivot_table(
            index=["lecture", "cond_id"], columns="bin_label", values="highlight_sec", aggfunc="sum", fill_value=0.0
        ).reset_index()

    # ============================================================
    # Write Excel
    # ============================================================
    readme = pd.DataFrame(
        [
            ["Inputs", "experiments/runs/<run_id>/extracted/index.csv + extracted/<lecture>/<cond>/meta.json"],
            ["Animation clips", "output_root/lecture_outputs/add_animation_outputs/slide_###_sent##(_style).mp4"],
            ["Final outputs", "output_root/lecture_outputs/output_final (lecture_final.mp4 等を探索)"],
            ["Highlight definition", "anim_type(=style) が非空の clip をハイライトとしてカウント"],
            ["Bins", f"abs bins: {bin_sec}s, pct bins: {pct_bins}"],
            ["Note", "分布（時間ビン）は clip 連結タイムライン上で算出。最終 mp4 があっても分布は clip を用いる。"],
        ],
        columns=["Topic", "Notes"],
    )

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        if df_style is not None and not df_style.empty:
            df_style.to_excel(writer, sheet_name="StyleDist", index=False)
        if df_bbox is not None and not df_bbox.empty:
            df_bbox.to_excel(writer, sheet_name="BBoxDist", index=False)
        if df_slide is not None and not df_slide.empty:
            df_slide.to_excel(writer, sheet_name="SlideDist", index=False)

        if df_bin60 is not None and not df_bin60.empty:
            df_bin60.to_excel(writer, sheet_name=f"Bins_{bin_sec}s", index=False)
        if df_binpct is not None and not df_binpct.empty:
            df_binpct.to_excel(writer, sheet_name=f"Bins_{int(100/pct_bins)}pct", index=False)

        if df_heat60 is not None:
            df_heat60.to_excel(writer, sheet_name="Heatmap_Abs", index=False)
        if df_heatpct is not None:
            df_heatpct.to_excel(writer, sheet_name="Heatmap_Pct", index=False)

        if df_raw is not None and not df_raw.empty:
            df_raw.to_excel(writer, sheet_name="Raw_Clips", index=False)

        wb = writer.book

        # format tweaks
        for name in wb.sheetnames:
            ws = wb[name]
            ws.freeze_panes = "A2"
            autosize_columns(ws)

        # heatmap conditional formatting
        if "Heatmap_Abs" in wb.sheetnames:
            ws = wb["Heatmap_Abs"]
            # data starts at C2 (A=lecture, B=cond_id)
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row >= 2 and max_col >= 3:
                add_heatmap(ws, "C2", f"{get_column_letter(max_col)}{max_row}")

        if "Heatmap_Pct" in wb.sheetnames:
            ws = wb["Heatmap_Pct"]
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row >= 2 and max_col >= 3:
                add_heatmap(ws, "C2", f"{get_column_letter(max_col)}{max_row}")

    log(f"wrote: {out_xlsx}")
    return out_xlsx


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-id", required=True)
    ap.add_argument("--runs-root", default=str(RUNS_ROOT))
    ap.add_argument("--bin-sec", type=int, default=60)
    ap.add_argument("--pct-bins", type=int, default=20)
    ap.add_argument("--allow-partial", action="store_true")
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()

    out = build(
        run_id=args.run_id,
        runs_root=Path(args.runs_root),
        bin_sec=args.bin_sec,
        pct_bins=args.pct_bins,
        allow_partial=bool(args.allow_partial),
        force=bool(args.force),
    )
    print(str(out))


if __name__ == "__main__":
    main()