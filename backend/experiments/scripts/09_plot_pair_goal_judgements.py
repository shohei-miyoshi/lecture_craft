# experiments/scripts/09_plot_pair_goal_judgements.py
# -*- coding: utf-8 -*-
"""
Step9: 卒論貼り付け用 Excel を生成する（実験1〜4）

- 実験1: Step4 metrics_basic.csv から形式指標の増減率（教材ごとの増減率の平均）
- 実験2: Step5 gpt_counts.csv から内容指標を集計
        - 増減率%（教材ごとの増減率の平均）※左=0は除外（発散回避）
        - ★正規化密度（1000字あたり）の条件別平均（NA不要）
        - ★正規化密度の差分（cond - base）の比較別平均（NA不要）
- 実験3: Step6 pair_judgements.csv + JSON から勝率（overall + GPT-4o/GPT-5）
        ※ Step8 を介さず Step6 を直接集計（向きの違いで落ちない）
- 実験4: Step7 animation_review.csv（あれば）を貼り付け（なければテンプレ）

出力:
  experiments/runs/<run_id>/analysis/step9_thesis/step9_thesis.xlsx

実行:
  python experiments/scripts/09_plot_pair_goal_judgements.py --run-id <run_id> --force
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Paths
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parents[2]
RUNS_ROOT = PROJECT_ROOT / "experiments" / "runs"
TEMPLATE_CONFIG_DEFAULT = PROJECT_ROOT / "experiments" / "config" / "experiment_config.json"


# ============================================================
# Labels
# ============================================================
KEY_LABEL_JA = {
    "summary": "要約らしさ（大枠・簡潔）",
    "detail": "詳細らしさ（掘り下げ）",
    "intro": "入門らしさ（易しい・前提補完）",
    "advanced": "発展らしさ（高度・専門）",
    "modality_audio_fitness": "音声提示の適合性",
    "modality_visual_fitness": "動画提示の適合性",
    "goal_overview": "目的: 大枠を知りたい",
    "goal_quick": "目的: 手短に知りたい",
    "goal_deep_listening": "目的: じっくり聞きたい",
}

MODEL_DISPLAY = {
    "gpt-4o": "GPT-4o",
    "gpt-5": "GPT-5",
}

COND_LABEL_JA = {
    "baseline_anim": "動画ベースライン",
    "summary_only_anim": "動画要約",
    "detail_only_anim": "動画詳細",
    "intro_only_anim": "動画入門",
    "advanced_only_anim": "動画発展",
    "baseline_audio": "音声ベースライン",
    "intro_only_audio": "音声入門",
    "summary_only_audio": "音声要約",
    "combo_adv_detail_anim": "動画combo（発展+詳細）",
    "combo_intro_summary_audio": "音声combo（入門+要約）",
}

# 実験1（形式）で使う指標（Step4の列名）
E1_METRICS = [
    ("script_char_len", "台本文字数"),
    ("sentence_count", "文数"),
    ("avg_sentence_len", "平均文長"),
    ("max_sentence_len", "最大文長"),
]

# 実験2（内容）で使う指標（Step5 aspect_id）
E2_ASPECTS = [
    ("example_count", "例示"),
    ("causal_count", "根拠・理由づけ"),
    ("procedure_count", "手順"),
    ("definition_count", "定義"),
    ("assumption_count", "前提補完"),
    ("technical_term_count", "専門用語"),
    ("visual_reference_count", "視覚参照"),
    ("deictic_count", "指示語"),
    ("paraphrase_count", "言い換え"),
    ("attention_count", "注意喚起"),
    ("vague_count", "曖昧表現"),
]


# ============================================================
# Utilities
# ============================================================
def read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def safe_float(x: Any, default: float = 0.0) -> float:
    try:
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def is_baseline(cond_id: str) -> bool:
    c = (cond_id or "").strip()
    return c.startswith("baseline_") or c in ("baseline_anim", "baseline_audio")


def canonical_pair(a: str, b: str) -> Tuple[str, str, str]:
    """
    Step8と同じ規則で anchor/other/comparison_id を作る。
    - baseline が絡むなら baseline を anchor
    - baseline が無いなら辞書順で anchor 固定（向き問題を根絶）
    """
    A = (a or "").strip()
    B = (b or "").strip()
    if is_baseline(A) and not is_baseline(B):
        anchor, other = A, B
    elif is_baseline(B) and not is_baseline(A):
        anchor, other = B, A
    else:
        anchor, other = (A, B) if A <= B else (B, A)
    return anchor, other, f"{anchor}__vs__{other}"


def cond_label(cond_id: str) -> str:
    return COND_LABEL_JA.get(cond_id, cond_id)


def key_label(key: str) -> str:
    return KEY_LABEL_JA.get(key, key)


def model_label(model: str) -> str:
    return MODEL_DISPLAY.get(model, model)


def pct_fmt_cell_value(x: Optional[float]) -> Any:
    # Excel表示を "80.0%" のようにしたい（0.8ではなく80.0を入れる）
    return "" if x is None else float(x)


# ============================================================
# Excel writer helpers
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

PCT_NUMBER_FORMAT = '0.0"%"'
NUM_NUMBER_FORMAT = "0.0"
DENS_NUMBER_FORMAT = "0.00"  # 密度・密度差分


def write_table(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    headers: List[str],
    rows: List[List[Any]],
    freeze_header: bool = True,
    pct_cols: Optional[List[int]] = None,  # 0-based within table
    num_cols: Optional[List[int]] = None,  # 0-based within table
    col_formats: Optional[Dict[int, str]] = None,  # 0-based within table
) -> int:
    """
    戻り値: 書き終えた最終行（1-indexed）
    """
    pct_cols = pct_cols or []
    num_cols = num_cols or []
    col_formats = col_formats or {}
    r0 = start_row
    c0 = start_col

    # header
    for j, h in enumerate(headers):
        cell = ws.cell(row=r0, column=c0 + j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER

    # body
    for i, row in enumerate(rows, start=1):
        rr = r0 + i
        for j, v in enumerate(row):
            cell = ws.cell(row=rr, column=c0 + j, value=v)
            cell.border = BORDER
            cell.alignment = ALIGN_LEFT if j == 0 else ALIGN_CENTER

            if j in col_formats and v != "":
                cell.number_format = col_formats[j]
            elif j in pct_cols and v != "":
                cell.number_format = PCT_NUMBER_FORMAT
            elif j in num_cols and v != "":
                cell.number_format = NUM_NUMBER_FORMAT

    if freeze_header:
        ws.freeze_panes = ws.cell(row=r0 + 1, column=c0)

    # auto width (simple)
    max_col = c0 + len(headers) - 1
    for col in range(c0, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for rr in range(r0, r0 + 1 + len(rows)):
            v = ws.cell(row=rr, column=col).value
            s = "" if v is None else str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    return r0 + len(rows)


def add_block_title(ws: Worksheet, row: int, col: int, title: str) -> int:
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = Font(bold=True, size=12)
    ws.row_dimensions[row].height = 20
    return row + 1


# ============================================================
# Run config
# ============================================================
@dataclass(frozen=True)
class CondSpec:
    type: str     # "animation" or "audio"
    level: str    # "baseline"/"intro"/"advanced"
    detail: str   # "baseline"/"summary"/"detail"


def load_run_config(run_dir: Path) -> Dict[str, Any]:
    snap = run_dir / "config" / "experiment_config.json"
    if snap.exists():
        return read_json(snap)
    return read_json(TEMPLATE_CONFIG_DEFAULT)


def load_cond_specs(cfg: Dict[str, Any]) -> Dict[str, CondSpec]:
    out: Dict[str, CondSpec] = {}
    specs = cfg.get("condition_specs", {})
    for k, v in specs.items():
        out[k] = CondSpec(type=v["type"], level=v["level"], detail=v["detail"])
    return out


# ============================================================
# Experiment 1 (Step4)
# ============================================================
def _pick_lecture_id(row: Dict[str, str]) -> str:
    for k in ("lecture_key", "lecture_title", "lecture", "material", "deck"):
        v = (row.get(k) or "").strip()
        if v:
            return v
    return ""


def load_step4_metrics(run_dir: Path) -> Dict[Tuple[str, str], Dict[str, float]]:
    """
    (lecture_id, cond_id) -> metrics dict
    lecture_id は lecture_key を優先（無い場合は lecture_title 等）
    """
    p = run_dir / "analysis" / "metrics_basic" / "metrics_basic.csv"
    if not p.exists():
        return {}
    rows = read_csv(p)
    out: Dict[Tuple[str, str], Dict[str, float]] = {}
    for r in rows:
        lec = _pick_lecture_id(r)
        cond = (r.get("cond_id") or "").strip()
        if not lec or not cond:
            continue
        d: Dict[str, float] = {}
        for k, _ja in E1_METRICS:
            d[k] = safe_float(r.get(k), default=0.0)
        out[(lec, cond)] = d
    return out


def mean_ratio_percent(
    table: Dict[Tuple[str, str], Dict[str, float]],
    left_cond: str,
    right_cond: str,
    metric_key: str,
) -> Tuple[Optional[float], int]:
    """
    mean over lectures of (right-left)/left*100
    NOTE: left==0 は除外（発散回避）
    """
    ratios: List[float] = []
    lectures = sorted({lec for (lec, _c) in table.keys()})
    for lec in lectures:
        L = table.get((lec, left_cond))
        R = table.get((lec, right_cond))
        if not L or not R:
            continue
        denom = float(L.get(metric_key, 0.0))
        if denom == 0.0:
            continue
        ratios.append((float(R.get(metric_key, 0.0)) - float(L.get(metric_key, 0.0))) / denom * 100.0)
    if not ratios:
        return None, 0
    return sum(ratios) / float(len(ratios)), len(ratios)


# ============================================================
# Experiment 2 (Step5)
# ============================================================
def load_step5_counts(run_dir: Path) -> Dict[Tuple[str, str, str], float]:
    """
    (lecture_key, cond_id, aspect_id) -> value
    """
    p = run_dir / "analysis" / "gpt_counts.csv"
    if not p.exists():
        return {}
    rows = read_csv(p)
    out: Dict[Tuple[str, str, str], float] = {}
    for r in rows:
        ok = (r.get("ok") or "").strip()
        if ok != "1":
            continue
        lec = (r.get("lecture_key") or "").strip()
        cond = (r.get("cond_id") or "").strip()
        aspect = (r.get("aspect_id") or "").strip()
        if not lec or not cond or not aspect:
            continue
        out[(lec, cond, aspect)] = safe_float(r.get("value"), default=0.0)
    return out


def mean_ratio_percent_step5(
    table: Dict[Tuple[str, str, str], float],
    left_cond: str,
    right_cond: str,
    aspect_id: str,
) -> Tuple[Optional[float], int]:
    """
    mean over lectures of (right-left)/left*100
    NOTE: left==0 は除外（発散回避）
    """
    ratios: List[float] = []
    lectures = sorted({lec for (lec, _c, _a) in table.keys()})
    for lec in lectures:
        L = table.get((lec, left_cond, aspect_id))
        R = table.get((lec, right_cond, aspect_id))
        if L is None or R is None:
            continue
        denom = float(L)
        if denom == 0.0:
            continue
        ratios.append((float(R) - float(L)) / denom * 100.0)
    if not ratios:
        return None, 0
    return sum(ratios) / float(len(ratios)), len(ratios)


def build_e2_table_ratio(
    counts: Dict[Tuple[str, str, str], float],
    comparisons: List[Tuple[str, str, str]],  # (label, left, right)
) -> List[List[Any]]:
    headers = ["比較", "n"] + [f"{ja} 増減率" for _k, ja in E2_ASPECTS]
    out: List[List[Any]] = []
    for label, left, right in comparisons:
        row: List[Any] = [label, ""]
        ns: List[int] = []
        for aspect, _ja in E2_ASPECTS:
            mean_pct, n = mean_ratio_percent_step5(counts, left, right, aspect)
            row.append(pct_fmt_cell_value(mean_pct))
            ns.append(n)
        row[1] = max(ns) if ns else 0
        out.append(row)
    return [headers, *out]


def build_e2_norm_by_conditions(
    counts: Dict[Tuple[str, str, str], float],
    step4_metrics: Dict[Tuple[str, str], Dict[str, float]],
    cond_ids: List[str],
    scale: float = 1000.0,
) -> Tuple[List[str], List[List[Any]]]:
    """
    卒論向けの「条件別 正規化平均」表。

    - 正規化: density(u) = count(u) / script_char_len(u) * scale
    - 平均: 教材 u ごとの density(u) を平均
    - n: その条件で「script_char_len が取得でき、かつ全観点が揃う教材数」
    """
    headers = ["条件", "n"] + [f"{ja}（/1000字）" for _a, ja in E2_ASPECTS]
    lectures_all = sorted({lec for (lec, _c, _a) in counts.keys()})

    rows: List[List[Any]] = []
    for cond in cond_ids:
        lectures_ok: List[str] = []
        for lec in lectures_all:
            m = step4_metrics.get((lec, cond))
            if not m:
                continue
            char_len = float(m.get("script_char_len", 0.0))
            if char_len <= 0.0:
                continue
            ok = True
            for aspect, _ja in E2_ASPECTS:
                if (lec, cond, aspect) not in counts:
                    ok = False
                    break
            if ok:
                lectures_ok.append(lec)

        n = len(lectures_ok)
        row: List[Any] = [cond_label(cond), n]

        for aspect, _ja in E2_ASPECTS:
            if n == 0:
                row.append("")
                continue
            vals: List[float] = []
            for lec in lectures_ok:
                cnt = float(counts[(lec, cond, aspect)])
                char_len = float(step4_metrics[(lec, cond)]["script_char_len"])
                vals.append(cnt / char_len * scale)
            row.append(sum(vals) / float(len(vals)))

        rows.append(row)

    return headers, rows


def _density_one(
    counts: Dict[Tuple[str, str, str], float],
    step4_metrics: Dict[Tuple[str, str], Dict[str, float]],
    lecture_key: str,
    cond_id: str,
    aspect_id: str,
    scale: float = 1000.0,
) -> Optional[float]:
    m = step4_metrics.get((lecture_key, cond_id))
    if not m:
        return None
    char_len = float(m.get("script_char_len", 0.0))
    if char_len <= 0.0:
        return None
    v = counts.get((lecture_key, cond_id, aspect_id))
    if v is None:
        return None
    return float(v) / char_len * scale


def mean_density_delta_step5(
    counts: Dict[Tuple[str, str, str], float],
    step4_metrics: Dict[Tuple[str, str], Dict[str, float]],
    left_cond: str,
    right_cond: str,
    aspect_id: str,
    scale: float = 1000.0,
) -> Tuple[Optional[float], int]:
    """
    mean over lectures of (density_right - density_left)
    NOTE: 0問題がなく、NAを作らずに比較できる（両方取れる教材のみ平均）
    """
    deltas: List[float] = []
    lectures_all = sorted({lec for (lec, _c, _a) in counts.keys()})
    for lec in lectures_all:
        dl = _density_one(counts, step4_metrics, lec, left_cond, aspect_id, scale=scale)
        dr = _density_one(counts, step4_metrics, lec, right_cond, aspect_id, scale=scale)
        if dl is None or dr is None:
            continue
        deltas.append(dr - dl)
    if not deltas:
        return None, 0
    return sum(deltas) / float(len(deltas)), len(deltas)


def build_e2_table_density_delta(
    counts: Dict[Tuple[str, str, str], float],
    step4_metrics: Dict[Tuple[str, str], Dict[str, float]],
    comparisons: List[Tuple[str, str, str]],  # (label, left, right)
    scale: float = 1000.0,
) -> Tuple[List[str], List[List[Any]]]:
    """
    比較ごとの「密度差分（right-left）」表（/1000字）。NAは出さない（差分は常に定義可能）。
    """
    headers = ["比較", "n"] + [f"{ja} Δ（/1000字）" for _a, ja in E2_ASPECTS]
    rows: List[List[Any]] = []
    for label, left, right in comparisons:
        row: List[Any] = [label, ""]
        ns: List[int] = []
        for aspect, _ja in E2_ASPECTS:
            mean_delta, n = mean_density_delta_step5(
                counts, step4_metrics, left, right, aspect, scale=scale
            )
            row.append("" if mean_delta is None else float(mean_delta))
            ns.append(n)
        row[1] = max(ns) if ns else 0
        rows.append(row)
    return headers, rows


# ============================================================
# Experiment 3 (Step6 direct aggregation)
# ============================================================
@dataclass(frozen=True)
class JudgementRow:
    model: str
    lecture: str
    pair_id: str
    A_cond: str
    B_cond: str
    key: str
    winner_raw: str          # "A"/"B"/"tie"/"unknown"
    winner_cond: str         # cond_id or "tie"/"unknown"
    anchor: str
    other: str
    comparison_id: str


def winner_ab_to_cond(winner_raw: str, A_cond: str, B_cond: str) -> str:
    w = (winner_raw or "").strip()
    if w == "A":
        return (A_cond or "").strip()
    if w == "B":
        return (B_cond or "").strip()
    if w == "tie":
        return "tie"
    return "unknown"


def load_step6_judgements(run_dir: Path) -> List[JudgementRow]:
    """
    Step6:
      analysis/pair_judgements.csv
      analysis/pair_judgements/<model>/<lecture>/<pair_id>.json
    JSONの judgements を正として展開する（CSVの keys は信用しない）
    """
    p = run_dir / "analysis" / "pair_judgements.csv"
    if not p.exists():
        return []

    rows = read_csv(p)
    out: List[JudgementRow] = []

    for r in rows:
        ok = (r.get("ok") or "").strip()
        if ok != "1":
            continue
        json_path = (r.get("json_path") or "").strip()
        if not json_path:
            continue
        jp = Path(json_path)
        if not jp.is_absolute():
            jp = (run_dir / jp).resolve()
        if not jp.exists():
            continue

        model = (r.get("model") or "").strip()
        lecture = (r.get("lecture") or "").strip()
        pair_id = (r.get("pair_id") or "").strip()
        A_cond = (r.get("A_cond") or "").strip()
        B_cond = (r.get("B_cond") or "").strip()
        if not model or not lecture or not pair_id or not A_cond or not B_cond:
            continue

        payload = read_json(jp)
        jud = payload.get("judgements") or {}
        if not isinstance(jud, dict):
            continue

        anchor, other, comp_id = canonical_pair(A_cond, B_cond)

        for key, obj in jud.items():
            if not isinstance(obj, dict):
                continue
            winner_raw = (obj.get("winner") or "").strip()
            if winner_raw == "":
                winner_raw = "unknown"
            winner_cond = winner_ab_to_cond(winner_raw, A_cond, B_cond)

            out.append(JudgementRow(
                model=model,
                lecture=lecture,
                pair_id=pair_id,
                A_cond=A_cond,
                B_cond=B_cond,
                key=str(key),
                winner_raw=winner_raw,
                winner_cond=winner_cond,
                anchor=anchor,
                other=other,
                comparison_id=comp_id,
            ))
    return out


def target_cond_for_key(
    key: str,
    cond1: str,
    cond2: str,
    specs: Dict[str, CondSpec],
) -> str:
    a = specs.get(cond1)
    b = specs.get(cond2)
    if a is None or b is None:
        if cond1.startswith("combo_"):
            return cond1
        if cond2.startswith("combo_"):
            return cond2
        return cond2

    if key == "summary":
        if a.detail == "summary":
            return cond1
        if b.detail == "summary":
            return cond2
        return cond2

    if key == "detail":
        if a.detail == "detail":
            return cond1
        if b.detail == "detail":
            return cond2
        return cond2

    if key == "intro":
        if a.level == "intro":
            return cond1
        if b.level == "intro":
            return cond2
        return cond2

    if key == "advanced":
        if a.level == "advanced":
            return cond1
        if b.level == "advanced":
            return cond2
        return cond2

    if key == "modality_audio_fitness":
        if a.type == "audio":
            return cond1
        if b.type == "audio":
            return cond2
        return cond2

    if key == "modality_visual_fitness":
        if a.type == "animation":
            return cond1
        if b.type == "animation":
            return cond2
        return cond2

    if key.startswith("goal_"):
        if cond1.startswith("combo_"):
            return cond1
        if cond2.startswith("combo_"):
            return cond2
        return cond2

    if cond1.startswith("combo_"):
        return cond1
    if cond2.startswith("combo_"):
        return cond2
    return cond2


@dataclass(frozen=True)
class WinAgg:
    win: int
    lose: int
    tie: int
    n: int

    @property
    def win_rate_percent(self) -> Optional[float]:
        if self.n <= 0:
            return None
        return (float(self.win) / float(self.n)) * 100.0


def aggregate_wins(
    judgements: List[JudgementRow],
    comparison_id: str,
    key: str,
    target_cond: str,
    model_filter: Optional[str] = None,
) -> WinAgg:
    win = lose = tie = 0
    for r in judgements:
        if r.comparison_id != comparison_id:
            continue
        if r.key != key:
            continue
        if model_filter is not None and r.model != model_filter:
            continue

        if r.winner_cond == "unknown":
            continue

        if r.winner_cond == "tie":
            tie += 1
        elif r.winner_cond == target_cond:
            win += 1
        else:
            lose += 1

    n = win + lose + tie
    return WinAgg(win=win, lose=lose, tie=tie, n=n)


# ============================================================
# Definitions: which comparisons to show (Experiment 3 sheets)
# ============================================================
@dataclass(frozen=True)
class CompDef:
    item: str
    label: str
    cond_a: str
    cond_b: str
    keys: List[str]


def build_comp_defs() -> Tuple[List[CompDef], List[CompDef], List[CompDef]]:
    e3_main = [
        CompDef("1-1", "1-1 動画baseline → 動画要約", "baseline_anim", "summary_only_anim", ["summary"]),
        CompDef("1-2", "1-2 動画baseline → 動画詳細", "baseline_anim", "detail_only_anim", ["detail"]),
        CompDef("2-1", "2-1 動画baseline → 動画入門", "baseline_anim", "intro_only_anim", ["intro"]),
        CompDef("2-2", "2-2 動画baseline → 動画発展", "baseline_anim", "advanced_only_anim", ["advanced"]),
        CompDef("3-2", "3-2 モダリティ: 動画baseline ↔ 音声baseline", "baseline_anim", "baseline_audio",
                ["modality_audio_fitness", "modality_visual_fitness"]),
    ]

    e3_audio = [
        CompDef("4A-1", "4A-1 音声combo: 入門単独 vs combo（summary観点）",
                "intro_only_audio", "combo_intro_summary_audio", ["summary"]),
        CompDef("4A-2", "4A-2 音声combo: 要約単独 vs combo（intro観点）",
                "summary_only_audio", "combo_intro_summary_audio", ["intro"]),
        CompDef("4A-3", "4A-3 音声combo: 音声baseline → combo（summary+intro+overview）",
                "baseline_audio", "combo_intro_summary_audio", ["goal_overview", "intro", "summary"]),
        CompDef("4A-4", "4A-4 音声combo: 動画baseline → combo（summary+intro+quick+modality）",
                "baseline_anim", "combo_intro_summary_audio",
                ["goal_quick", "intro", "summary", "modality_audio_fitness", "modality_visual_fitness"]),
    ]

    e3_video = [
        CompDef("4V-1", "4V-1 動画combo: 詳細単独 vs combo（advanced観点）",
                "detail_only_anim", "combo_adv_detail_anim", ["advanced"]),
        CompDef("4V-2", "4V-2 動画combo: 発展単独 vs combo（detail観点）",
                "advanced_only_anim", "combo_adv_detail_anim", ["detail"]),
        CompDef("4V-3", "4V-3 動画combo: 動画baseline → combo（advanced+detail+deep_listening）",
                "baseline_anim", "combo_adv_detail_anim", ["goal_deep_listening", "advanced", "detail"]),
    ]
    return e3_main, e3_audio, e3_video


# ============================================================
# Build Experiment 3 tables (with per-model columns)
# ============================================================
def build_e3_table_rows(
    judgements: List[JudgementRow],
    specs: Dict[str, CondSpec],
    comp_defs: List[CompDef],
    models: List[str],
) -> List[List[Any]]:
    model_cols = ["gpt-4o", "gpt-5"]
    _ = models

    out_rows: List[List[Any]] = []

    for c in comp_defs:
        anchor, other, comp_id = canonical_pair(c.cond_a, c.cond_b)

        for key in c.keys:
            target = target_cond_for_key(key, anchor, other, specs)

            overall = aggregate_wins(judgements, comp_id, key, target, model_filter=None)

            per_model_rates: Dict[str, Optional[float]] = {}
            for m in model_cols:
                agg_m = aggregate_wins(judgements, comp_id, key, target, model_filter=m)
                per_model_rates[m] = agg_m.win_rate_percent

            out_rows.append([
                f"{c.item} {c.label} / {key_label(key)}",
                f"{c.item} {c.label}",
                key_label(key),
                cond_label(target),
                pct_fmt_cell_value(overall.win_rate_percent),
                pct_fmt_cell_value(per_model_rates.get("gpt-4o")),
                pct_fmt_cell_value(per_model_rates.get("gpt-5")),
                overall.win,
                overall.lose,
                overall.tie,
                overall.n,
            ])

    return out_rows


def build_e3_chart_block(
    e3_rows: List[List[Any]],
) -> Tuple[List[str], List[List[Any]]]:
    comps = []
    keys = []
    for r in e3_rows:
        comp = str(r[1])
        key = str(r[2])
        if comp not in comps:
            comps.append(comp)
        if key not in keys:
            keys.append(key)

    headers = ["比較"] + keys
    mat: List[List[Any]] = []
    lookup: Dict[Tuple[str, str], Any] = {}
    for r in e3_rows:
        lookup[(str(r[1]), str(r[2]))] = r[4]  # overall win rate

    for comp in comps:
        row = [comp]
        for key in keys:
            v = lookup.get((comp, key), "")
            row.append(v)
        mat.append(row)

    return headers, mat


# ============================================================
# Build Experiment 1 tables
# ============================================================
def build_e1_table(
    metrics: Dict[Tuple[str, str], Dict[str, float]],
    comparisons: List[Tuple[str, str, str]],  # (label, left, right)
) -> List[List[Any]]:
    headers = ["比較", "n"] + [f"{ja} 増減率" for _k, ja in E1_METRICS]
    out: List[List[Any]] = []
    for label, left, right in comparisons:
        row: List[Any] = [label, ""]
        ns: List[int] = []
        for k, _ja in E1_METRICS:
            mean_pct, n = mean_ratio_percent(metrics, left, right, k)
            row.append(pct_fmt_cell_value(mean_pct))
            ns.append(n)
        row[1] = max(ns) if ns else 0
        out.append(row)
    return [headers, *out]


# ============================================================
# Experiment 4 table (Step7)
# ============================================================
def load_or_make_e4_rows(run_dir: Path) -> Tuple[List[str], List[List[Any]]]:
    p = run_dir / "review" / "animation_review" / "animation_review.csv"
    headers = [
        "lecture", "cond_id", "page", "sentence_no", "text",
        "対象一致(1-5)", "タイミング(1-5)", "過不足(1-5)", "妨害性(1-5)", "メモ",
    ]

    if p.exists():
        rows = read_csv(p)
        out: List[List[Any]] = []
        for r in rows:
            out.append([
                r.get("lecture", ""),
                r.get("cond_id", ""),
                r.get("page", ""),
                r.get("sentence_no", ""),
                r.get("text", ""),
                r.get("match_score", ""),
                r.get("timing_score", ""),
                r.get("amount_score", ""),
                r.get("disturb_score", ""),
                r.get("note", ""),
            ])
        return headers, out

    out = []
    for _ in range(10):
        out.append(["", "", "", "", "", "", "", "", "", ""])
    return headers, out


# ============================================================
# Main
# ============================================================
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-id", required=True)
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()

    run_dir = RUNS_ROOT / args.run_id
    if not run_dir.exists():
        raise FileNotFoundError(f"run_dir not found: {run_dir}")

    cfg = load_run_config(run_dir)
    specs = load_cond_specs(cfg)
    judge_models = cfg.get("judge_models", [])
    judge_models = [str(m).strip() for m in judge_models if str(m).strip()]

    # load inputs
    step4 = load_step4_metrics(run_dir)
    step5 = load_step5_counts(run_dir)
    step6 = load_step6_judgements(run_dir)

    # output excel
    out_dir = run_dir / "analysis" / "step9_thesis"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / "step9_thesis.xlsx"
    if out_xlsx.exists() and args.force:
        out_xlsx.unlink(missing_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------
    # README
    # ------------------------
    ws = wb.create_sheet("README")
    lines = [
        "このExcelは卒論貼り付け用の表と、範囲選択→挿入で図化できるブロックを含みます。",
        "",
        "[実験1] E1_RQ1-3a / E1_RQ4 : 形式指標（Step4 metrics_basic）",
        "[実験2] E2_RQ1-3a / E2_RQ4 : 内容指標（Step5 gpt_counts）",
        "          - 増減率%：比較元が0の教材は除外（発散回避）",
        "          - 正規化密度：count / script_char_len × 1000 を教材平均（条件別表）",
        "          - 正規化密度差分：密度(right) - 密度(left) を教材平均（比較別表）※NA不要",
        "[実験3] E3_RQ1-3a / E3_RQ4_audio / E3_RQ4_video : 勝率（Step6直接集計）",
        "          - 狙い側勝率(GPT-4o), 狙い側勝率(GPT-5) 列あり",
        "[実験4] E4_highlight : 人手評価（Step7, 無ければテンプレ）",
    ]
    for i, s in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=s)
    ws.column_dimensions["A"].width = 120

    # ------------------------
    # CONFIG
    # ------------------------
    ws = wb.create_sheet("CONFIG")
    cfg_rows: List[List[Any]] = []
    for cond_id, sp in specs.items():
        cfg_rows.append([cond_id, cond_label(cond_id), sp.type, sp.level, sp.detail])
    cfg_rows.sort(key=lambda x: str(x[0]))
    write_table(
        ws, 1, 1,
        ["cond_id", "日本語ラベル", "type", "level", "detail"],
        cfg_rows,
        pct_cols=[],
        num_cols=[],
    )

    # ------------------------
    # E1 tables
    # ------------------------
    ws = wb.create_sheet("E1_RQ1-3a")
    r = 1
    r = add_block_title(ws, r, 1, "実験1（形式）RQ1〜RQ3a：台本が“本当に変わった”証拠（増減率%）")
    e1_main_comps = [
        ("動画baseline → 動画要約", "baseline_anim", "summary_only_anim"),
        ("動画baseline → 動画詳細", "baseline_anim", "detail_only_anim"),
        ("動画baseline → 動画入門", "baseline_anim", "intro_only_anim"),
        ("動画baseline → 動画発展", "baseline_anim", "advanced_only_anim"),
        ("動画baseline → 音声baseline（モダリティ差）", "baseline_anim", "baseline_audio"),
    ]
    e1_table = build_e1_table(step4, e1_main_comps)
    headers = e1_table[0]
    rows = e1_table[1:]
    r_end = write_table(ws, r, 1, headers, rows, pct_cols=list(range(2, 2 + len(E1_METRICS))))
    r = r_end + 2

    ws = wb.create_sheet("E1_RQ4")
    r = 1
    r = add_block_title(ws, r, 1, "実験1（形式）RQ4：combo 比較（増減率%）")
    e1_rq4_comps = [
        ("動画combo: 詳細単独 → combo（発展+詳細）", "detail_only_anim", "combo_adv_detail_anim"),
        ("動画combo: 発展単独 → combo（発展+詳細）", "advanced_only_anim", "combo_adv_detail_anim"),
        ("動画combo: 動画baseline → combo（発展+詳細）", "baseline_anim", "combo_adv_detail_anim"),
        ("音声combo: 入門単独 → combo（入門+要約）", "intro_only_audio", "combo_intro_summary_audio"),
        ("音声combo: 要約単独 → combo（入門+要約）", "summary_only_audio", "combo_intro_summary_audio"),
        ("音声combo: 音声baseline → combo（入門+要約）", "baseline_audio", "combo_intro_summary_audio"),
        ("音声combo: 動画baseline → combo（入門+要約）", "baseline_anim", "combo_intro_summary_audio"),
    ]
    e1_table2 = build_e1_table(step4, e1_rq4_comps)
    headers = e1_table2[0]
    rows = e1_table2[1:]
    write_table(ws, r, 1, headers, rows, pct_cols=list(range(2, 2 + len(E1_METRICS))))

    # ------------------------
    # E2 tables  ★ここ：密度表 + 密度差分表（NA不要）
    # ------------------------
    ws = wb.create_sheet("E2_RQ1-3a")
    r = 1
    r = add_block_title(ws, r, 1, "実験2（内容）RQ1〜RQ3a：変化の“中身”（増減率%）")
    e2_main_comps = [
        ("動画baseline → 動画要約", "baseline_anim", "summary_only_anim"),
        ("動画baseline → 動画詳細", "baseline_anim", "detail_only_anim"),
        ("動画baseline → 動画入門", "baseline_anim", "intro_only_anim"),
        ("動画baseline → 動画発展", "baseline_anim", "advanced_only_anim"),
        ("動画baseline → 音声baseline（モダリティ差）", "baseline_anim", "baseline_audio"),
    ]
    e2_ratio = build_e2_table_ratio(step5, e2_main_comps)
    headers = e2_ratio[0]
    rows = e2_ratio[1:]
    pct_cols = list(range(2, 2 + len(E2_ASPECTS)))
    r_end = write_table(ws, r, 1, headers, rows, pct_cols=pct_cols)
    r = r_end + 2

    # 追加(A)：正規化密度（条件別）
    r = add_block_title(ws, r, 1, "（追加A）正規化密度の平均（1000字あたり）— 条件別")
    e2_norm_conds_main = [
        "baseline_anim",
        "summary_only_anim",
        "detail_only_anim",
        "intro_only_anim",
        "advanced_only_anim",
        "baseline_audio",
    ]
    norm_headers, norm_rows = build_e2_norm_by_conditions(step5, step4, e2_norm_conds_main, scale=1000.0)
    dens_formats = {j: DENS_NUMBER_FORMAT for j in range(2, len(norm_headers))}
    r_end2 = write_table(ws, r, 1, norm_headers, norm_rows, col_formats=dens_formats)
    r = r_end2 + 2

    # 追加(B)：正規化密度の差分（比較別） ※NA不要
    r = add_block_title(ws, r, 1, "（追加B）正規化密度の差分（cond - base の平均, /1000字）— 比較別（NA不要）")
    dd_headers, dd_rows = build_e2_table_density_delta(step5, step4, e2_main_comps, scale=1000.0)
    dd_formats = {j: DENS_NUMBER_FORMAT for j in range(2, len(dd_headers))}
    write_table(ws, r, 1, dd_headers, dd_rows, col_formats=dd_formats)

    ws = wb.create_sheet("E2_RQ4")
    r = 1
    r = add_block_title(ws, r, 1, "実験2（内容）RQ4：combo 比較（増減率%）")
    e2_rq4_comps = [
        ("動画combo: 詳細単独 → combo（発展+詳細）", "detail_only_anim", "combo_adv_detail_anim"),
        ("動画combo: 発展単独 → combo（発展+詳細）", "advanced_only_anim", "combo_adv_detail_anim"),
        ("動画combo: 動画baseline → combo（発展+詳細）", "baseline_anim", "combo_adv_detail_anim"),
        ("音声combo: 入門単独 → combo（入門+要約）", "intro_only_audio", "combo_intro_summary_audio"),
        ("音声combo: 要約単独 → combo（入門+要約）", "summary_only_audio", "combo_intro_summary_audio"),
        ("音声combo: 音声baseline → combo（入門+要約）", "baseline_audio", "combo_intro_summary_audio"),
        ("音声combo: 動画baseline → combo（入門+要約）", "baseline_anim", "combo_intro_summary_audio"),
    ]
    e2_ratio2 = build_e2_table_ratio(step5, e2_rq4_comps)
    headers = e2_ratio2[0]
    rows = e2_ratio2[1:]
    pct_cols = list(range(2, 2 + len(E2_ASPECTS)))
    r_end = write_table(ws, r, 1, headers, rows, pct_cols=pct_cols)
    r = r_end + 2

    # 追加(A)：正規化密度（RQ4：動画combo系）
    r = add_block_title(ws, r, 1, "（追加A）正規化密度の平均：動画combo系（1000字あたり）")
    e2_norm_conds_video_combo = [
        "baseline_anim",
        "detail_only_anim",
        "advanced_only_anim",
        "combo_adv_detail_anim",
    ]
    norm_headers2, norm_rows2 = build_e2_norm_by_conditions(step5, step4, e2_norm_conds_video_combo, scale=1000.0)
    dens_formats2 = {j: DENS_NUMBER_FORMAT for j in range(2, len(norm_headers2))}
    r_end2 = write_table(ws, r, 1, norm_headers2, norm_rows2, col_formats=dens_formats2)
    r = r_end2 + 2

    # 追加(A)：正規化密度（RQ4：音声combo系）
    r = add_block_title(ws, r, 1, "（追加A）正規化密度の平均：音声combo系（1000字あたり）")
    e2_norm_conds_audio_combo = [
        "baseline_audio",
        "intro_only_audio",
        "summary_only_audio",
        "combo_intro_summary_audio",
        "baseline_anim",  # 参照（不要なら削除可）
    ]
    norm_headers3, norm_rows3 = build_e2_norm_by_conditions(step5, step4, e2_norm_conds_audio_combo, scale=1000.0)
    dens_formats3 = {j: DENS_NUMBER_FORMAT for j in range(2, len(norm_headers3))}
    r_end3 = write_table(ws, r, 1, norm_headers3, norm_rows3, col_formats=dens_formats3)
    r = r_end3 + 2

    # 追加(B)：正規化密度の差分（RQ4 比較別）※NA不要
    r = add_block_title(ws, r, 1, "（追加B）正規化密度の差分（cond - base の平均, /1000字）— RQ4 比較別（NA不要）")
    dd_headers2, dd_rows2 = build_e2_table_density_delta(step5, step4, e2_rq4_comps, scale=1000.0)
    dd_formats2 = {j: DENS_NUMBER_FORMAT for j in range(2, len(dd_headers2))}
    write_table(ws, r, 1, dd_headers2, dd_rows2, col_formats=dd_formats2)

    # ------------------------
    # E3 tables (3 sheets)
    # ------------------------
    e3_main_defs, e3_audio_defs, e3_video_defs = build_comp_defs()

    def write_e3_sheet(sheet_name: str, title: str, comp_defs: List[CompDef]) -> None:
        ws = wb.create_sheet(sheet_name)
        r = 1
        r = add_block_title(ws, r, 1, title)

        e3_rows = build_e3_table_rows(step6, specs, comp_defs, judge_models)

        headers = [
            "項目", "比較", "観点", "狙い側条件",
            "狙い側勝率",
            "狙い側勝率(GPT-4o)",
            "狙い側勝率(GPT-5)",
            "狙い側勝", "狙い側負", "引分", "n",
        ]
        pct_cols = [4, 5, 6]
        r_end = write_table(ws, r, 1, headers, e3_rows, pct_cols=pct_cols)

        r2 = r_end + 2
        r2 = add_block_title(ws, r2, 1, "（作図用）比較×観点の勝率（overall）ブロック：範囲選択→挿入で棒グラフ化")
        ch_headers, ch_rows = build_e3_chart_block(e3_rows)
        write_table(ws, r2, 1, ch_headers, ch_rows, pct_cols=list(range(1, len(ch_headers))))

    write_e3_sheet(
        "E3_RQ1-3a",
        "実験3（相対評価）RQ1〜RQ3a：狙い特徴を持つ側の勝率（Step6直接集計）",
        e3_main_defs,
    )
    write_e3_sheet(
        "E3_RQ4_audio",
        "実験3（相対評価）RQ4：音声combo の勝率（Step6直接集計）",
        e3_audio_defs,
    )
    write_e3_sheet(
        "E3_RQ4_video",
        "実験3（相対評価）RQ4：動画combo の勝率（Step6直接集計）",
        e3_video_defs,
    )

    # ------------------------
    # E4 highlight
    # ------------------------
    ws = wb.create_sheet("E4_highlight")
    r = 1
    r = add_block_title(ws, r, 1, "実験4：ハイライト妥当性（人手評価）")
    e4_headers, e4_rows = load_or_make_e4_rows(run_dir)
    write_table(ws, r, 1, e4_headers, e4_rows, pct_cols=[], num_cols=[])

    wb.save(out_xlsx)
    print(f"[Step9] wrote: {out_xlsx}")


if __name__ == "__main__":
    main()
